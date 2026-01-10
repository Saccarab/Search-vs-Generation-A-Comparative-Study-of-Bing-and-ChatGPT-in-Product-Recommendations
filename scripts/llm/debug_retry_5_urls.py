"""
Debug helper: retry a small set of URLs against Gemini and print raw responses/errors.

Reads `geo_updated.xlsx` to find `content_path` for each URL, fills the structured prompt
from prompts/page_label_prompt_v1.txt, then calls Gemini.

Usage:
  python scripts/llm/debug_retry_5_urls.py --xlsx geo_updated.xlsx --urls "u1,u2,u3"
"""

from __future__ import annotations

import argparse
import json
import os
import re
import time
from pathlib import Path
from typing import Dict, Optional, Tuple

import openpyxl
import requests


GEMINI_ENDPOINT_TMPL = "https://generativelanguage.googleapis.com/v1beta/{model_path}:generateContent?key={key}"


def safe_str(v: object) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() == "nan" else s


def _gemini_model_path(model: str) -> str:
    m = (model or "").strip()
    if not m:
        return "models/gemini-flash-latest"
    return m if m.startswith("models/") else f"models/{m}"


def sheet_header_map(ws) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        k = str(v).strip()
        if k:
            m[k] = c
    return m


def extract_url_domain(url: str) -> str:
    try:
        from urllib.parse import urlparse

        u = url.strip()
        if "://" not in u:
            u = "https://" + u
        host = (urlparse(u).netloc or "").lower()
        return host[4:] if host.startswith("www.") else host
    except Exception:
        return ""


def build_best_title_snippet(wb) -> Dict[str, Tuple[str, str]]:
    out: Dict[str, Tuple[str, str]] = {}
    if "bing_results" in wb.sheetnames:
        ws = wb["bing_results"]
        h = sheet_header_map(ws)
        if all(k in h for k in ("url", "result_title", "snippet")):
            for r in range(2, ws.max_row + 1):
                u = safe_str(ws.cell(r, h["url"]).value)
                if not u or u in out:
                    continue
                t = safe_str(ws.cell(r, h["result_title"]).value)
                s = safe_str(ws.cell(r, h["snippet"]).value)
                if t or s:
                    out[u] = (t, s)
    if "citations" in wb.sheetnames:
        ws = wb["citations"]
        h = sheet_header_map(ws)
        if all(k in h for k in ("url", "citation_title")):
            for r in range(2, ws.max_row + 1):
                u = safe_str(ws.cell(r, h["url"]).value)
                if not u or u in out:
                    continue
                t = safe_str(ws.cell(r, h["citation_title"]).value)
                if t:
                    out[u] = (t, "")
    return out


def fill_prompt(template: str, *, url: str, title: str, snippet: str, text: str) -> str:
    dom = extract_url_domain(url)
    return (
        template.replace("{URL}", url)
        .replace("{URL_DOMAIN}", dom)
        .replace("{TITLE}", title or "")
        .replace("{SNIPPET_OR_META_DESCRIPTION}", snippet or "")
        .replace("{EXTRACTED_TEXT}", (text or "")[:12000])
    )


def gemini_call_raw(*, api_key: str, model: str, prompt: str, timeout: int) -> Tuple[int, str]:
    payload = {"contents": [{"role": "user", "parts": [{"text": prompt}]}], "generationConfig": {"temperature": 0.0}}
    ep = GEMINI_ENDPOINT_TMPL.format(model_path=_gemini_model_path(model), key=api_key)
    r = requests.post(ep, json=payload, timeout=timeout)
    return r.status_code, r.text


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="geo_updated.xlsx")
    ap.add_argument("--prompt-file", default="prompts/page_label_prompt_v1.txt")
    ap.add_argument("--model", default="gemini-3-flash-preview")
    ap.add_argument("--timeout", type=int, default=25, help="Per-request timeout seconds (default: 25)")
    ap.add_argument("--urls", required=True, help="Comma-separated URLs to retry")
    args = ap.parse_args()

    api_key = (
        os.getenv("GEMINI_API_KEY", "").strip()
        or os.getenv("GOOGLE_API_KEY", "").strip()
        or os.getenv("GOOGLE_GENERATIVE_AI_API_KEY", "").strip()
    )
    if not api_key:
        raise SystemExit("Missing GEMINI_API_KEY in env.")

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    ws = wb["urls"]
    h = sheet_header_map(ws)
    if "url" not in h or "content_path" not in h:
        raise SystemExit("urls sheet missing url/content_path columns")

    prompt_template = Path(args.prompt_file).read_text(encoding="utf-8")
    lookup = build_best_title_snippet(wb)

    # Build url->content_path map
    cp_map: Dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        u = safe_str(ws.cell(r, h["url"]).value)
        if not u or u in cp_map:
            continue
        cp_map[u] = safe_str(ws.cell(r, h["content_path"]).value)

    urls = [safe_str(u) for u in args.urls.split(",") if safe_str(u)]
    for url in urls:
        print("\n=== URL ===", flush=True)
        print(url, flush=True)
        cp = cp_map.get(url, "")
        if not cp or not os.path.exists(cp):
            print("NO content_path file found:", cp, flush=True)
            continue
        text = Path(cp).read_text(encoding="utf-8", errors="ignore")
        title, snippet = lookup.get(url, ("", ""))
        prompt = fill_prompt(prompt_template, url=url, title=title, snippet=snippet, text=text)

        try:
            t0 = time.time()
            print(f"Calling Gemini (timeout={args.timeout}s)...", flush=True)
            status, body = gemini_call_raw(api_key=api_key, model=args.model, prompt=prompt, timeout=args.timeout)
            dt_s = time.time() - t0
            print("HTTP:", status, f"({dt_s:.1f}s)", flush=True)
            print("Body snippet:", (body or "")[:1200], flush=True)
            if status == 200:
                # Attempt parse into the same JSON we expect
                try:
                    data = json.loads(body)
                    out_text = data["candidates"][0]["content"]["parts"][0].get("text", "")
                    out_text = re.sub(r"^```(?:json)?\\s*|\\s*```$", "", (out_text or "").strip(), flags=re.I | re.M)
                    parsed = json.loads(out_text)
                    print("Parsed keys:", list(parsed.keys()) if isinstance(parsed, dict) else type(parsed), flush=True)
                except Exception as e:
                    print("Parse error:", type(e).__name__, str(e)[:300], flush=True)
        except Exception as e:
            print("Request error:", type(e).__name__, str(e)[:600], flush=True)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

