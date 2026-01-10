"""
Ingest Bing extension export (with full content extraction) into geo_updated.xlsx.

Why this exists:
- The extension export CSV can contain large per-URL extracted text (`content`).
- We do NOT store raw page text in Excel. Instead we write it to disk and store a `content_path`
  pointer in the `urls` sheet, along with a few deterministic fields we can compute from it.

Input CSV expectations (best-effort; missing columns are tolerated):
  - url (required)
  - content, contentLength, contentError (optional but recommended)
  - has_schema_markup (optional)
  - canonical_url / page_title / meta_description / published_date / modified_date (optional; ignored unless workbook has columns)

Writes to geo_updated.xlsx sheet: `urls`
  - url (key)
  - domain
  - content_path (relative path under content-dir)
  - content_word_count
  - has_schema_markup
  - fetched_at

Usage:
  python scripts/ingest/ingest_bing_content_export_to_geo_xlsx.py \
    --xlsx geo_updated.xlsx \
    --bing-export "C:/Users/User/Downloads/bing_results_2026-01-08T....csv" \
    --content-dir data/url_content
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import re
from pathlib import Path
from typing import Dict, Optional
from urllib.parse import urlparse

import openpyxl
import pandas as pd


def utc_now_iso() -> str:
    return dt.datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def safe_str(v: object) -> str:
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except Exception:
        pass
    s = str(v).strip()
    if s.lower() == "nan":
        return ""
    return s


def extract_domain(raw_url: str) -> str:
    u = safe_str(raw_url)
    if not u:
        return ""
    if "://" not in u:
        u = "https://" + u
    try:
        host = (urlparse(u).netloc or "").lower()
        return host[4:] if host.startswith("www.") else host
    except Exception:
        return ""


WORD_RE = re.compile(r"[A-Za-z0-9]+(?:'[A-Za-z0-9]+)?")


def count_words(text: str) -> int:
    if not text:
        return 0
    return len(WORD_RE.findall(text))


def short_hash(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()[:16]


def sheet_header_map(ws) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        k = str(v).strip()
        if k:
            m[k] = c
    return m


def find_row_by_key(ws, key_col_idx: int, key_val: str) -> Optional[int]:
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=key_col_idx).value
        if v is None:
            continue
        if str(v).strip() == key_val:
            return r
    return None


def append_row_first_truly_empty(ws, header: Dict[str, int], row: Dict[str, object]) -> int:
    # Find first truly empty row (templates often have preformatted blanks far down).
    r = None
    for rr in range(2, ws.max_row + 1):
        has_any = False
        for col_idx in header.values():
            v = ws.cell(row=rr, column=col_idx).value
            if v is not None and str(v).strip() != "":
                has_any = True
                break
        if not has_any:
            r = rr
            break
    if r is None:
        r = ws.max_row + 1
    for k, v in row.items():
        if k in header:
            ws.cell(row=r, column=header[k]).value = v
    return r


def upsert_row(ws, header: Dict[str, int], key_col: str, key_val: str, row: Dict[str, object]) -> int:
    key_idx = header[key_col]
    existing = find_row_by_key(ws, key_idx, key_val)
    if existing is None:
        row2 = dict(row)
        row2[key_col] = key_val
        return append_row_first_truly_empty(ws, header, row2)
    for k, v in row.items():
        if k not in header:
            continue
        if v is None:
            continue
        if isinstance(v, str) and v.strip() == "":
            continue
        ws.cell(row=existing, column=header[k]).value = v
    return existing


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Path to geo_updated.xlsx")
    ap.add_argument("--bing-export", required=True, help="Path to raw Bing extension export CSV (with content columns)")
    ap.add_argument("--content-dir", default="data/url_content", help="Directory to store extracted content files")
    ap.add_argument("--overwrite-content", action="store_true", help="Overwrite existing content files if present")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    csv_path = Path(args.bing_export)
    content_dir = Path(args.content_dir)
    content_dir.mkdir(parents=True, exist_ok=True)

    df = pd.read_csv(csv_path, engine="python")
    if "url" not in df.columns:
        raise SystemExit("CSV missing required column: url")

    wb = openpyxl.load_workbook(xlsx_path)
    ws_urls = wb["urls"]
    h_urls = sheet_header_map(ws_urls)

    required = ["url", "domain", "content_path", "content_word_count", "has_schema_markup", "fetched_at"]
    missing = [c for c in required if c not in h_urls]
    if missing:
        raise SystemExit(f"geo_updated.xlsx urls sheet missing required columns: {missing}")

    total_rows = 0
    unique_urls = set()
    wrote_content = 0
    updated_urls = 0
    skipped_no_content = 0

    for _, r in df.iterrows():
        raw_url = safe_str(r.get("url"))
        if not raw_url:
            continue
        total_rows += 1
        unique_urls.add(raw_url)

        content = safe_str(r.get("content"))
        content_err = safe_str(r.get("contentError"))
        # contentLength may be present but don't trust it; measure ourselves.
        has_schema = r.get("has_schema_markup")

        # Decide if we consider content usable
        usable = bool(content) and not content_err and len(content) >= 200
        content_path = ""
        word_count = ""
        if usable:
            # stable-ish filename per URL
            fname = f"{short_hash(raw_url)}.txt"
            # Use absolute output path for writing
            out_path = content_dir / fname
            if args.overwrite_content or not out_path.exists():
                out_path.write_text(content, encoding="utf-8", errors="ignore")
                wrote_content += 1
            content_path = str(Path(args.content_dir) / fname).replace("\\", "/")
            word_count = str(count_words(content))
        else:
            skipped_no_content += 1

        # Map schema flag
        hs = ""
        if isinstance(has_schema, bool):
            hs = "1" if has_schema else "0"
        else:
            hs_s = safe_str(has_schema).lower()
            if hs_s in ("1", "true", "yes", "y"):
                hs = "1"
            elif hs_s in ("0", "false", "no", "n"):
                hs = "0"

        row = {
            "domain": extract_domain(raw_url),
            "content_path": content_path,
            "content_word_count": word_count,
            "has_schema_markup": hs,
            "fetched_at": utc_now_iso(),
        }
        upsert_row(ws_urls, h_urls, "url", raw_url, row)
        updated_urls += 1

    wb.save(xlsx_path)

    print("Ingest complete.")
    print(f"Rows scanned: {total_rows}")
    print(f"Unique URLs: {len(unique_urls)}")
    print(f"URLs updated: {updated_urls}")
    print(f"Content files written: {wrote_content} (overwrite={'yes' if args.overwrite_content else 'no'})")
    print(f"URLs with unusable/empty content: {skipped_no_content}")


if __name__ == "__main__":
    main()

