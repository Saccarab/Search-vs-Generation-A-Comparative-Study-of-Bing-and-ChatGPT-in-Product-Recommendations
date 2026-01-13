"""
Run Gemini labeling on geo_updated.xlsx using the project's structured prompt:
  prompts/page_label_prompt_v1.txt

This reads URLs from `urls` where `content_path` exists, sends the prompt with
{URL}/{TITLE}/{SNIPPET_OR_META_DESCRIPTION}/{EXTRACTED_TEXT} filled, and writes:
  - `urls` labels (type/content_format/tone/etc.)
  - `listicles` row when type == listicle
  - `listicle_products` rows when type == listicle

Idempotent by default: skips rows that already have `urls.type` unless --overwrite.

Env:
  - GEMINI_API_KEY (preferred) OR GOOGLE_API_KEY OR GOOGLE_GENERATIVE_AI_API_KEY

Usage (Windows):
  set GEMINI_API_KEY=...
  python scripts/llm/enrich_geo_urls_with_gemini.py --xlsx geo_updated.xlsx --max-urls 50
"""

from __future__ import annotations

import argparse
import json
import os
import re
import time
import random
import sys
import hashlib
from pathlib import Path
from typing import Dict, Optional, Set, Tuple

import openpyxl
import requests


def safe_str(v: object) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() == "nan" else s


def sheet_header_map(ws) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        k = str(v).strip()
        if k:
            m[k] = c
    return m


def _gemini_model_path(model: str) -> str:
    m = (model or "").strip()
    if not m:
        return "models/gemini-flash-latest"
    return m if m.startswith("models/") else f"models/{m}"


GEMINI_ENDPOINT_TMPL = "https://generativelanguage.googleapis.com/v1beta/{model_path}:generateContent?key={key}"


def gemini_generate_json(*, api_key: str, model: str, prompt: str, timeout: int = 60) -> dict:
    payload = {"contents": [{"role": "user", "parts": [{"text": prompt}]}], "generationConfig": {"temperature": 0.0}}
    model_path = _gemini_model_path(model)
    ep = GEMINI_ENDPOINT_TMPL.format(model_path=model_path, key=api_key)
    r = requests.post(ep, json=payload, timeout=timeout)
    # Keep a short snippet around for debugging if we fail.
    resp_text_snippet = (r.text or "")[:800]
    r.raise_for_status()
    data = r.json()
    out_text = data["candidates"][0]["content"]["parts"][0].get("text", "")
    out_text = re.sub(r"^```(?:json)?\\s*|\\s*```$", "", (out_text or "").strip(), flags=re.IGNORECASE | re.MULTILINE)
    try:
        return json.loads(out_text)
    except json.JSONDecodeError as e:
        # Provide actionable context to the caller/log.
        raise ValueError(f"Gemini returned non-JSON. error={e}. raw_text_snippet={(out_text or '')[:800]!r}. http_body_snippet={resp_text_snippet!r}") from e


def sheet_header_map(ws) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        k = str(v).strip()
        if k:
            m[k] = c
    return m


def find_row_by_key(ws, col_idx: int, key_val: str) -> Optional[int]:
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, col_idx).value
        if v is None:
            continue
        if str(v).strip() == key_val:
            return r
    return None


def append_row_first_truly_empty(ws, header: Dict[str, int], row: Dict[str, object]) -> int:
    for rr in range(2, ws.max_row + 1):
        has_any = False
        for col_idx in header.values():
            v = ws.cell(rr, col_idx).value
            if v is not None and str(v).strip() != "":
                has_any = True
                break
        if not has_any:
            for k, v in row.items():
                if k in header:
                    ws.cell(rr, header[k]).value = v
            return rr
    rr = ws.max_row + 1
    for k, v in row.items():
        if k in header:
            ws.cell(rr, header[k]).value = v
    return rr


def upsert_row(ws, header: Dict[str, int], key_col: str, key_val: str, row: Dict[str, object], overwrite: bool) -> int:
    key_idx = header[key_col]
    existing = find_row_by_key(ws, key_idx, key_val)
    if existing is None:
        row2 = dict(row)
        row2[key_col] = key_val
        return append_row_first_truly_empty(ws, header, row2)
    for k, v in row.items():
        if k not in header:
            continue
        if v is None:
            continue
        if isinstance(v, str) and v.strip() == "":
            continue
        if not overwrite:
            cur = ws.cell(existing, header[k]).value
            if cur is not None and str(cur).strip() != "":
                continue
        ws.cell(existing, header[k]).value = v
    return existing


def upsert_listicle_product_row(
    ws_products,
    hp: Dict[str, int],
    *,
    listicle_url: str,
    position_in_listicle: str,
    product_name: str,
    item: Dict[str, object],
    overwrite: bool,
) -> int:
    """
    Upsert listicle_products by composite key (listicle_url, position_in_listicle, product_name).
    If overwrite is True, update existing row's non-empty fields.
    """
    key_cols = (hp["listicle_url"], hp["position_in_listicle"], hp["product_name"])
    for rr in range(2, ws_products.max_row + 1):
        a = ws_products.cell(rr, key_cols[0]).value
        b = ws_products.cell(rr, key_cols[1]).value
        c = ws_products.cell(rr, key_cols[2]).value
        if a is None or b is None or c is None:
            continue
        if (str(a).strip(), str(b).strip(), str(c).strip()) != (listicle_url, position_in_listicle, product_name):
            continue
        # found existing
        if overwrite:
            for k, v in item.items():
                if k not in hp:
                    continue
                if v is None:
                    continue
                if isinstance(v, str) and v.strip() == "":
                    continue
                ws_products.cell(rr, hp[k]).value = v
        return rr
    return append_row_first_truly_empty(ws_products, hp, item)


def extract_url_domain(url: str) -> str:
    try:
        from urllib.parse import urlparse

        u = url.strip()
        if "://" not in u:
            u = "https://" + u
        host = (urlparse(u).netloc or "").lower()
        return host[4:] if host.startswith("www.") else host
    except Exception:
        return ""


def build_best_title_snippet(wb) -> Dict[str, Tuple[str, str]]:
    """
    Best-effort mapping url -> (title, snippet/meta).
    Preference:
      - bing_results: result_title + snippet
      - citations: citation_title + empty snippet
    """
    out: Dict[str, Tuple[str, str]] = {}

    if "bing_results" in wb.sheetnames:
        ws = wb["bing_results"]
        h = sheet_header_map(ws)
        if all(k in h for k in ("url", "result_title", "snippet")):
            for r in range(2, ws.max_row + 1):
                u = safe_str(ws.cell(r, h["url"]).value)
                if not u or u in out:
                    continue
                t = safe_str(ws.cell(r, h["result_title"]).value)
                s = safe_str(ws.cell(r, h["snippet"]).value)
                if t or s:
                    out[u] = (t, s)

    if "citations" in wb.sheetnames:
        ws = wb["citations"]
        h = sheet_header_map(ws)
        if all(k in h for k in ("url", "citation_title")):
            for r in range(2, ws.max_row + 1):
                u = safe_str(ws.cell(r, h["url"]).value)
                if not u or u in out:
                    continue
                t = safe_str(ws.cell(r, h["citation_title"]).value)
                if t:
                    out[u] = (t, "")

    return out


def fill_prompt(template: str, *, url: str, title: str, snippet: str, text: str) -> str:
    dom = extract_url_domain(url)
    return (
        template.replace("{URL}", url)
        .replace("{URL_DOMAIN}", dom)
        .replace("{TITLE}", title or "")
        .replace("{SNIPPET_OR_META_DESCRIPTION}", snippet or "")
        .replace("{EXTRACTED_TEXT}", (text or "")[:12000])
    )

def build_gemini_payload(prompt: str) -> dict:
    return {"contents": [{"role": "user", "parts": [{"text": prompt}]}], "generationConfig": {"temperature": 0.0}}


def approx_tokens_from_chars(n_chars: int) -> int:
    # crude heuristic for English-ish text: ~4 chars/token
    return int(max(0, n_chars) / 4)


def dump_payload(*, dump_dir: str, url: str, prompt: str, model: str) -> str:
    """
    Write the exact request payload body (no API key) to disk for deterministic replay.
    Returns the written file path.
    """
    os.makedirs(dump_dir, exist_ok=True)
    h = hashlib.sha256(url.encode("utf-8")).hexdigest()[:12]
    out = os.path.join(dump_dir, f"payload_{h}.json")
    payload = build_gemini_payload(prompt)
    payload["_debug"] = {"url": url, "model": model, "prompt_chars": len(prompt), "approx_tokens": approx_tokens_from_chars(len(prompt))}
    with open(out, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False)
    return out

def load_processed_urls_from_jsonl(path: str) -> Tuple[Set[str], Set[str]]:
    """
    Return (ok_urls, failed_urls) from a JSONL audit log.
    We use this to resume runs without re-calling the API if Excel hasn't been saved yet.
    """
    ok: Set[str] = set()
    fail: Set[str] = set()
    if not path or not os.path.exists(path):
        return ok, fail
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except Exception:
                continue
            if not isinstance(obj, dict):
                continue
            u = safe_str(obj.get("url"))
            if not u:
                continue
            if obj.get("ok") is True:
                ok.add(u)
            elif obj.get("ok") is False:
                fail.add(u)
    return ok, fail

def _load_only_urls(only_urls: str, only_urls_file: str) -> Set[str]:
    out: Set[str] = set()
    if only_urls:
        for part in only_urls.split(","):
            u = safe_str(part)
            if u:
                out.add(u)
    if only_urls_file:
        p = Path(only_urls_file)
        if p.exists():
            for line in p.read_text(encoding="utf-8", errors="ignore").splitlines():
                u = safe_str(line)
                if u:
                    out.add(u)
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--max-urls", type=int, default=0, help="If >0, process at most N URLs")
    ap.add_argument("--model", default="gemini-1.5-flash-latest")
    ap.add_argument("--timeout", type=int, default=60)
    ap.add_argument("--overwrite", action="store_true", help="Overwrite existing labels if present")
    ap.add_argument("--prompt-file", default="prompts/page_label_prompt_v1.txt", help="Prompt template path")
    ap.add_argument("--sleep", type=float, default=0.0, help="Sleep seconds between API calls")
    ap.add_argument("--log-jsonl", default="data/llm/page_labels_gemini.jsonl", help="Append JSONL audit log")
    ap.add_argument("--save-every", type=int, default=20, help="Save geo_updated.xlsx every N processed URLs (0 = only at end)")
    ap.add_argument(
        "--retry-failures",
        action="store_true",
        help="If set, re-attempt URLs previously logged as ok=false in the JSONL audit log.",
    )
    ap.add_argument("--retries", type=int, default=2, help="Retry count for transient failures (ReadTimeout/429/503).")
    ap.add_argument("--retry-backoff", type=float, default=2.0, help="Base backoff seconds between retries.")
    ap.add_argument(
        "--retry-timeout-mult",
        type=float,
        default=1.5,
        help="Multiply timeout on each retry (e.g. 1.5 => 60s, 90s, 135s...).",
    )
    ap.add_argument(
        "--max-seconds-per-url",
        type=float,
        default=0.0,
        help="Hard budget per URL for Gemini attempts (seconds). 0 disables. "
        "Example: 45 means we will not spend more than ~45s total on a single URL (including retries/backoff).",
    )
    ap.add_argument(
        "--verbose",
        action="store_true",
        help="Print one line per API attempt (URL, attempt #, timeout, success/error, backoff).",
    )
    ap.add_argument(
        "--log-payload-stats",
        action="store_true",
        help="Print payload stats per URL (prompt chars, approx tokens).",
    )
    ap.add_argument(
        "--dump-payloads-dir",
        default="",
        help="If set, write the exact request payload JSON (no API key) for failed URLs into this directory.",
    )
    ap.add_argument(
        "--only-urls",
        default="",
        help="Optional: comma-separated list of URLs to process (debugging/targeted retry).",
    )
    ap.add_argument(
        "--only-urls-file",
        default="",
        help="Optional: text file with one URL per line to process (debugging/targeted retry).",
    )
    args = ap.parse_args()

    api_key = (
        os.getenv("GEMINI_API_KEY", "").strip()
        or os.getenv("GOOGLE_API_KEY", "").strip()
        or os.getenv("GOOGLE_GENERATIVE_AI_API_KEY", "").strip()
    )
    if not api_key:
        raise SystemExit("Missing API key env var. Set GEMINI_API_KEY (preferred).")

    wb = openpyxl.load_workbook(args.xlsx)
    ws_urls = wb["urls"]
    ws_listicles = wb["listicles"]
    ws_products = wb["listicle_products"]
    hu = sheet_header_map(ws_urls)
    hl = sheet_header_map(ws_listicles)
    hp = sheet_header_map(ws_products)

    required_urls = [
        "url",
        "domain",
        "content_path",
        "type",
        "content_format",
        "tone",
        "promotional_intensity_score",
        "freshness_cue_strength",
        "has_pros_cons",
        "has_clear_authorship",
        "has_sources_or_citations",
        "expertise_signal_score",
        "spamminess_score",
        "primary_intent",
    ]
    missing = [c for c in required_urls if c not in hu]
    if missing:
        raise SystemExit(f"urls sheet missing required columns: {missing}")
    if "listicle_url" not in hl:
        raise SystemExit("listicles sheet missing listicle_url column")
    if not all(k in hp for k in ("listicle_url", "product_name", "position_in_listicle")):
        raise SystemExit("listicle_products sheet missing required columns (listicle_url/product_name/position_in_listicle)")

    prompt_path = Path(args.prompt_file)
    template = prompt_path.read_text(encoding="utf-8")
    os.makedirs(Path(args.log_jsonl).parent, exist_ok=True)
    lookup = build_best_title_snippet(wb)
    ok_urls, fail_urls = load_processed_urls_from_jsonl(args.log_jsonl)
    only_urls = _load_only_urls(args.only_urls, args.only_urls_file)

    processed = 0
    labeled = 0
    skipped = 0
    failed = 0
    last_saved_at = 0

    for r in range(2, ws_urls.max_row + 1):
        url = safe_str(ws_urls.cell(r, hu["url"]).value)
        if not url:
            continue
        if only_urls and url not in only_urls:
            continue
        cp = safe_str(ws_urls.cell(r, hu["content_path"]).value)
        if not cp or not os.path.exists(cp):
            continue

        existing_type = safe_str(ws_urls.cell(r, hu["type"]).value)
        if not args.overwrite and existing_type:
            skipped += 1
            continue
        # Resume support: if we already logged this URL, skip it unless overwrite or retry-failures applies.
        if not args.overwrite:
            if url in ok_urls:
                skipped += 1
                continue
            if (not args.retry_failures) and (url in fail_urls):
                skipped += 1
                continue

        processed += 1
        if args.max_urls and processed > args.max_urls:
            break

        try:
            text = open(cp, "r", encoding="utf-8", errors="ignore").read()
            title, snippet = lookup.get(url, ("", ""))
            prompt = fill_prompt(template, url=url, title=title, snippet=snippet, text=text)
            if args.log_payload_stats:
                print(
                    f"[payload] url={url} prompt_chars={len(prompt)} approx_tokens~{approx_tokens_from_chars(len(prompt))}",
                    flush=True,
                )
            # Retry transient failures (timeouts / throttling) with backoff.
            parsed = None
            last_err: Optional[Exception] = None
            attempts = max(1, int(args.retries) + 1)
            url_t0 = time.time()
            for attempt in range(1, attempts + 1):
                try:
                    # Enforce per-URL time budget if set
                    if args.max_seconds_per_url and args.max_seconds_per_url > 0:
                        elapsed = time.time() - url_t0
                        remaining = float(args.max_seconds_per_url) - elapsed
                        if remaining <= 0:
                            raise TimeoutError(
                                f"Per-URL budget exceeded ({args.max_seconds_per_url}s) before attempt {attempt}/{attempts}"
                            )
                    else:
                        remaining = None

                    t = int(float(args.timeout) * (float(args.retry_timeout_mult) ** (attempt - 1)))
                    if remaining is not None:
                        # Don't start an attempt that can't possibly finish within budget
                        t = max(1, min(t, int(remaining)))
                    if args.verbose:
                        print(
                            f"[gemini] url={url} attempt={attempt}/{attempts} timeout={t}s",
                            flush=True,
                        )
                    t0 = time.time()
                    parsed = gemini_generate_json(api_key=api_key, model=args.model, prompt=prompt, timeout=t)
                    if args.verbose:
                        dt = time.time() - t0
                        print(f"[gemini] ok url={url} dt={dt:.1f}s", flush=True)
                    last_err = None
                    break
                except Exception as e:
                    last_err = e
                    # Decide if retryable
                    retryable = False
                    if isinstance(e, (requests.exceptions.ReadTimeout, requests.exceptions.ConnectTimeout, requests.exceptions.ConnectionError)):
                        retryable = True
                    elif isinstance(e, requests.HTTPError) and getattr(e, "response", None) is not None:
                        try:
                            sc = int(e.response.status_code)
                            if sc in (429, 500, 502, 503, 504):
                                retryable = True
                        except Exception:
                            pass

                    if (attempt >= attempts) or (not retryable):
                        if args.verbose:
                            print(
                                f"[gemini] fail url={url} attempt={attempt}/{attempts} retryable={retryable} err={type(e).__name__}: {str(e)[:200]}",
                                flush=True,
                            )
                        break
                    # backoff + jitter
                    base = max(0.0, float(args.retry_backoff))
                    sleep_s = base * (2 ** (attempt - 1)) + random.random() * 0.25
                    if args.max_seconds_per_url and args.max_seconds_per_url > 0:
                        elapsed = time.time() - url_t0
                        remaining = float(args.max_seconds_per_url) - elapsed
                        if remaining <= 0:
                            break
                        sleep_s = min(sleep_s, max(0.0, remaining))
                    if args.verbose:
                        print(
                            f"[gemini] retrying url={url} in {sleep_s:.2f}s (err={type(e).__name__})",
                            flush=True,
                        )
                    time.sleep(sleep_s)

            if parsed is None:
                raise last_err or RuntimeError("Gemini request failed without exception.")

            # --- urls ---
            uo = parsed.get("urls") if isinstance(parsed, dict) else None
            if isinstance(uo, dict):
                def _num(v, default=""):
                    try:
                        if v is None:
                            return default
                        if isinstance(v, bool):
                            return "1" if v else "0"
                        return str(int(float(v)))
                    except Exception:
                        return default

                ws_urls.cell(r, hu["type"]).value = safe_str(uo.get("type"))
                ws_urls.cell(r, hu["content_format"]).value = safe_str(uo.get("content_format"))
                ws_urls.cell(r, hu["tone"]).value = safe_str(uo.get("tone"))
                ws_urls.cell(r, hu["promotional_intensity_score"]).value = _num(uo.get("promotional_intensity_score"))
                ws_urls.cell(r, hu["freshness_cue_strength"]).value = _num(uo.get("freshness_cue_strength"))
                ws_urls.cell(r, hu["has_pros_cons"]).value = _num(uo.get("has_pros_cons"))
                ws_urls.cell(r, hu["has_clear_authorship"]).value = _num(uo.get("has_clear_authorship"))
                ws_urls.cell(r, hu["has_sources_or_citations"]).value = _num(uo.get("has_sources_or_citations"))
                ws_urls.cell(r, hu["expertise_signal_score"]).value = _num(uo.get("expertise_signal_score"))
                ws_urls.cell(r, hu["spamminess_score"]).value = _num(uo.get("spamminess_score"))
                ws_urls.cell(r, hu["primary_intent"]).value = safe_str(uo.get("primary_intent"))

            # --- listicles + products ---
            lo = parsed.get("listicles") if isinstance(parsed, dict) else None
            if isinstance(lo, dict):
                upsert_row(ws_listicles, hl, "listicle_url", url, lo, overwrite=args.overwrite)

            po = parsed.get("listicle_products") if isinstance(parsed, dict) else None
            if isinstance(po, list):
                for item in po:
                    if not isinstance(item, dict):
                        continue
                    lk = safe_str(item.get("listicle_url") or url)
                    pos = safe_str(item.get("position_in_listicle"))
                    pn = safe_str(item.get("product_name"))
                    if not lk or not pos or not pn:
                        continue
                    # Optional: product_url (per-item) if present in the sheet/prompt output
                    if "product_url" in hp:
                        item["product_url"] = safe_str(item.get("product_url"))
                    upsert_listicle_product_row(
                        ws_products,
                        hp,
                        listicle_url=lk,
                        position_in_listicle=pos,
                        product_name=pn,
                        item=item,
                        overwrite=args.overwrite,
                    )

            labeled += 1

            # Audit log (jsonl)
            with open(args.log_jsonl, "a", encoding="utf-8") as lf:
                lf.write(json.dumps({"url": url, "ok": True, "response": parsed}) + "\n")
            ok_urls.add(url)
        except Exception as e:
            failed += 1
            # Capture useful failure info for debugging + targeted retries.
            err = {
                "type": type(e).__name__,
                "message": str(e)[:1200],
            }
            # If it's an HTTP error, include status code + response snippet when available
            if isinstance(e, requests.HTTPError) and getattr(e, "response", None) is not None:
                try:
                    err["http_status"] = int(e.response.status_code)
                    err["http_body_snippet"] = (e.response.text or "")[:800]
                except Exception:
                    pass
            with open(args.log_jsonl, "a", encoding="utf-8") as lf:
                lf.write(json.dumps({"url": url, "ok": False, "error": err}) + "\n")
            fail_urls.add(url)
            if args.verbose:
                print(f"[gemini] wrote failure log url={url} type={err.get('type')}", flush=True)
            if args.dump_payloads_dir:
                try:
                    out = dump_payload(dump_dir=args.dump_payloads_dir, url=url, prompt=prompt, model=args.model)
                    if args.verbose or args.log_payload_stats:
                        print(f"[payload] dumped failed payload -> {out}", flush=True)
                except Exception as de:
                    if args.verbose:
                        print(f"[payload] dump failed url={url} err={type(de).__name__}: {str(de)[:200]}", flush=True)

        if processed % 20 == 0:
            print(f"processed {processed} (labeled={labeled}, skipped={skipped}, failed={failed})")
        if args.sleep:
            time.sleep(args.sleep)
        if args.save_every and args.save_every > 0 and (processed - last_saved_at) >= args.save_every:
            wb.save(args.xlsx)
            last_saved_at = processed
            print(f"checkpoint saved at processed={processed}")
            sys.stdout.flush()

    wb.save(args.xlsx)
    print("Done.")
    print("processed:", processed)
    print("labeled:", labeled)
    print("skipped:", skipped)
    print("failed:", failed)


if __name__ == "__main__":
    main()

