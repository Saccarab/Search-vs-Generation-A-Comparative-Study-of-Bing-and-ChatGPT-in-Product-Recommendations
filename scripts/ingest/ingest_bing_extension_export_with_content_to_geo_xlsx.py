"""
Ingest a *raw Bing extension export* (which may include extracted page content + metadata)
into geo_updated.xlsx, while writing page text to disk and storing a pointer in `urls.content_path`.

This is meant to be the bridge between:
  (1) Running the Chrome extension (Bing SERP + optional content extraction)
  (2) Your Excel-centric workflow (geo_updated.xlsx)

What it does:
  - Writes `bing_results` rows (SERP-level):
      run_id, bing_query, result_rank, result_title, url, result_domain, snippet, page_num, captured_at
  - Upserts `urls` rows (URL-level):
      url, domain, content_path, content_word_count, has_schema_markup, fetched_at
  - Stores extracted `content` (text) into a folder you choose (e.g. thesis folder), NOT inside Excel.

Input CSV expected columns (best-effort):
  - query, position, run_id, title, snippet, domain, url, page_num
  - content, contentError, has_schema_markup (optional)

Usage example (Windows):
  python scripts/ingest/ingest_bing_extension_export_with_content_to_geo_xlsx.py ^
    --xlsx geo_updated.xlsx ^
    --bing-export "C:\\Users\\User\\Downloads\\bing_results_2026-01-08T15-14-19.csv" ^
    --content-root "C:\\Users\\User\\Documents\\thesis\\bing_content"
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import re
from pathlib import Path
from typing import Dict, Optional, Set
from urllib.parse import parse_qsl, urlparse, urlsplit, urlunsplit, urlencode

import base64

import openpyxl
import pandas as pd


WORD_RE = re.compile(r"[A-Za-z0-9]+(?:'[A-Za-z0-9]+)?")
TRACKING_PARAM_PREFIXES = ("utm_",)
TRACKING_PARAMS_EXACT = {
    "gclid",
    "fbclid",
    "msclkid",
    "yclid",
    "mc_cid",
    "mc_eid",
    "igshid",
    "ref",
    "ref_src",
    "spm",
}


def utc_now_iso() -> str:
    return dt.datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def safe_str(v: object) -> str:
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except Exception:
        pass
    s = str(v).strip()
    if s.lower() == "nan":
        return ""
    return s


def count_words(text: str) -> int:
    if not text:
        return 0
    return len(WORD_RE.findall(text))


def short_hash(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()[:16]


def extract_domain(raw_url: str) -> str:
    u = safe_str(raw_url)
    if not u:
        return ""
    if "://" not in u:
        u = "https://" + u
    try:
        host = (urlparse(u).netloc or "").lower()
        return host[4:] if host.startswith("www.") else host
    except Exception:
        return ""

def normalize_url_key(raw_url: str) -> str:
    """
    Normalized URL key for deduping joins while keeping the original URL stored.
    - lower host, strip leading www
    - drop fragment
    - remove trailing slash (except root)
    - drop common tracking params (utm_*, msclkid, fbclid, ...)
    - keep other query params (conservative)
    """
    u = safe_str(raw_url)
    if not u:
        return ""
    if "://" not in u:
        u = "https://" + u
    p = urlsplit(u)
    host = (p.netloc or "").lower()
    if host.startswith("www."):
        host = host[4:]
    path = p.path or "/"
    if len(path) > 1 and path.endswith("/"):
        path = path[:-1]
    kept = []
    for k, v in parse_qsl(p.query, keep_blank_values=True):
        lk = k.lower()
        if any(lk.startswith(px) for px in TRACKING_PARAM_PREFIXES):
            continue
        if lk in TRACKING_PARAMS_EXACT:
            continue
        kept.append((k, v))
    q = urlencode(kept, doseq=True)
    # scheme removed for key; fragment dropped
    return urlunsplit(("", host, path, q, "")).lstrip("/")


def decode_bing_ck_url(raw_url: str) -> str:
    """
    Decode Bing click-tracking redirect URLs (bing.com/ck/a) to their real destination if possible.
    If we can't decode to a real http(s) URL, return empty string.
    """
    u = safe_str(raw_url)
    if not u:
        return ""
    try:
        parts = urlsplit(u if "://" in u else "https://" + u)
        host = (parts.netloc or "").lower()
        if host.startswith("www."):
            host = host[4:]
        if not host.endswith("bing.com") or not parts.path.startswith("/ck/a"):
            return u

        qs = dict(parse_qsl(parts.query, keep_blank_values=True))
        enc = safe_str(qs.get("u"))
        if not enc.startswith("a1"):
            return ""
        b64 = enc[2:]
        pad = "=" * ((4 - (len(b64) % 4)) % 4)
        try:
            decoded = base64.urlsafe_b64decode((b64 + pad).encode("utf-8")).decode("utf-8", errors="replace")
        except Exception:
            return ""
        return decoded if decoded.startswith("http") else ""
    except Exception:
        return ""


def build_urls_key_index(ws_urls, h_urls) -> Dict[str, str]:
    """
    Map normalized key -> stored url (first seen) for collapsing URL variants into a single row.
    """
    out: Dict[str, str] = {}
    if "url" not in h_urls:
        return out
    url_col = h_urls["url"]
    for r in range(2, ws_urls.max_row + 1):
        v = ws_urls.cell(r, url_col).value
        if not v:
            continue
        s = str(v).strip()
        if not s:
            continue
        k = normalize_url_key(s)
        if k and k not in out:
            out[k] = s
    return out

def sheet_header_map(ws) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        k = str(v).strip()
        if k:
            m[k] = c
    return m


def find_row_by_key(ws, key_col_idx: int, key_val: str) -> Optional[int]:
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=key_col_idx).value
        if v is None:
            continue
        if str(v).strip() == key_val:
            return r
    return None


def append_row_first_truly_empty(ws, header: Dict[str, int], row: Dict[str, object]) -> int:
    r = None
    for rr in range(2, ws.max_row + 1):
        has_any = False
        for col_idx in header.values():
            v = ws.cell(row=rr, column=col_idx).value
            if v is not None and str(v).strip() != "":
                has_any = True
                break
        if not has_any:
            r = rr
            break
    if r is None:
        r = ws.max_row + 1
    for k, v in row.items():
        if k in header:
            ws.cell(row=r, column=header[k]).value = v
    return r


def upsert_row(ws, header: Dict[str, int], key_col: str, key_val: str, row: Dict[str, object], overwrite: bool) -> int:
    key_idx = header[key_col]
    existing = find_row_by_key(ws, key_idx, key_val)
    if existing is None:
        row2 = dict(row)
        row2[key_col] = key_val
        return append_row_first_truly_empty(ws, header, row2)
    for k, v in row.items():
        if k not in header:
            continue
        if v is None:
            continue
        if isinstance(v, str) and v.strip() == "":
            continue
        if not overwrite:
            cur = ws.cell(row=existing, column=header[k]).value
            if cur is not None and str(cur).strip() != "":
                continue
        ws.cell(row=existing, column=header[k]).value = v
    return existing


def existing_bing_keys(ws_bing, h_bing) -> Set[str]:
    """
    Build a set of keys to avoid duplicates on rerun.
    Key: run_id|bing_query|result_rank|url
    """
    out: Set[str] = set()
    need = ["run_id", "bing_query", "result_rank", "url"]
    if any(k not in h_bing for k in need):
        return out
    for r in range(2, ws_bing.max_row + 1):
        run_id = ws_bing.cell(r, h_bing["run_id"]).value
        q = ws_bing.cell(r, h_bing["bing_query"]).value
        rank = ws_bing.cell(r, h_bing["result_rank"]).value
        url = ws_bing.cell(r, h_bing["url"]).value
        if not run_id or not q or not rank or not url:
            continue
        out.add(f"{str(run_id).strip()}|{str(q).strip()}|{str(rank).strip()}|{str(url).strip()}")
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Path to geo_updated.xlsx")
    ap.add_argument("--bing-export", required=True, help="Path to Bing extension export CSV")
    ap.add_argument("--content-root", required=True, help="Folder to write extracted content into (outside repo is fine)")
    ap.add_argument("--run-label", default="", help="Optional subfolder name under content-root (default: CSV filename stem)")
    ap.add_argument("--overwrite", action="store_true", help="Overwrite existing `urls` fields (content_path, wordcount, etc.)")
    ap.add_argument("--max-content-chars", type=int, default=0, help="If >0, truncate written content files to N chars")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    csv_path = Path(args.bing_export)
    run_label = args.run_label.strip() or csv_path.stem
    content_root = Path(args.content_root) / run_label
    content_root.mkdir(parents=True, exist_ok=True)

    df = pd.read_csv(csv_path, engine="python")
    if "url" not in df.columns:
        raise SystemExit("CSV missing required column: url")

    wb = openpyxl.load_workbook(xlsx_path)
    ws_bing = wb["bing_results"]
    ws_urls = wb["urls"]
    h_bing = sheet_header_map(ws_bing)
    h_urls = sheet_header_map(ws_urls)
    urls_key_index = build_urls_key_index(ws_urls, h_urls)

    # Validate minimal schema we intend to write
    for col in ["run_id", "bing_query", "result_rank", "result_title", "url", "result_domain", "snippet", "captured_at", "page_num"]:
        if col not in h_bing:
            raise SystemExit(f"geo_updated.xlsx bing_results sheet missing required column: {col}")
    for col in ["url", "domain", "content_path", "content_word_count", "has_schema_markup", "fetched_at"]:
        if col not in h_urls:
            raise SystemExit(f"geo_updated.xlsx urls sheet missing required column: {col}")

    seen_bing = existing_bing_keys(ws_bing, h_bing)
    now = utc_now_iso()

    n_rows = 0
    n_bing_added = 0
    n_urls_upserted = 0
    n_content_written = 0

    for _, r in df.iterrows():
        url_raw = safe_str(r.get("url"))
        url = decode_bing_ck_url(url_raw) if "bing.com/ck/a" in url_raw.lower() else url_raw
        if not url and url_raw:
            # Skip rows where URL is only a tracking redirect that we couldn't decode.
            continue
        if not url:
            continue
        n_rows += 1

        run_id = safe_str(r.get("run_id"))
        query = safe_str(r.get("query"))
        # CSV uses `position` (1..N)
        rank = safe_str(r.get("position"))
        title = safe_str(r.get("title"))
        snippet = safe_str(r.get("snippet"))
        domain = safe_str(r.get("domain")) or extract_domain(url)
        page_num = safe_str(r.get("page_num"))

        # ---- bing_results append (idempotent) ----
        bing_key = f"{run_id}|{query}|{rank}|{url}"
        if run_id and query and rank and bing_key not in seen_bing:
            append_row_first_truly_empty(
                ws_bing,
                h_bing,
                {
                    "run_id": run_id,
                    "bing_query": query,
                    "result_rank": int(float(rank)) if rank else "",
                    "result_title": title,
                    "url": url,
                    "result_domain": domain,
                    "snippet": snippet,
                    "page_type": "",
                    "captured_at": now,
                    "page_num": int(float(page_num)) if page_num else "",
                },
            )
            seen_bing.add(bing_key)
            n_bing_added += 1

        # ---- urls upsert + content file ----
        content = safe_str(r.get("content"))
        content_err = safe_str(r.get("contentError"))
        content_path = ""
        wc = ""

        if content and not content_err:
            if args.max_content_chars and len(content) > args.max_content_chars:
                content = content[: args.max_content_chars]
            fname = f"{short_hash(url)}.txt"
            out_path = content_root / fname
            out_path.write_text(content, encoding="utf-8", errors="ignore")
            n_content_written += 1
            content_path = str(out_path)
            wc = str(count_words(content))

        hs = safe_str(r.get("has_schema_markup")).lower()
        hs_out = ""
        if hs in ("1", "true", "yes", "y"):
            hs_out = "1"
        elif hs in ("0", "false", "no", "n"):
            hs_out = "0"

        # Collapse URL variants by normalized key (http/https, fragments, tracking params, trailing slash, etc.)
        k = normalize_url_key(url)
        key_url = urls_key_index.get(k) if k else None
        if not key_url:
            key_url = url
            if k:
                urls_key_index[k] = key_url

        upsert_row(
            ws_urls,
            h_urls,
            "url",
            key_url,
            {
                "domain": domain,
                "content_path": content_path,
                "content_word_count": wc,
                "has_schema_markup": hs_out,
                "fetched_at": now,
            },
            overwrite=args.overwrite,
        )
        n_urls_upserted += 1

    wb.save(xlsx_path)

    print("Done.")
    print(f"Rows scanned: {n_rows}")
    print(f"bing_results rows appended: {n_bing_added}")
    print(f"urls rows upserted: {n_urls_upserted}")
    print(f"content files written: {n_content_written}")
    print(f"content root: {content_root}")


if __name__ == "__main__":
    main()

