import openpyxl
from pathlib import Path

def clear_sheets(xlsx_path, sheets_to_clear):
    wb = openpyxl.load_workbook(xlsx_path)
    for sheet_name in sheets_to_clear:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Keep headers (row 1), delete everything else
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row)
            print(f"Cleared sheet: {sheet_name}")
        else:
            print(f"Sheet not found: {sheet_name}")
    wb.save(xlsx_path)
    print(f"Saved {xlsx_path}")

if __name__ == "__main__":
    xlsx = "geo-fresh.xlsx"
    sheets = ["runs", "citations", "bing_results", "urls"]
    clear_sheets(xlsx, sheets)
