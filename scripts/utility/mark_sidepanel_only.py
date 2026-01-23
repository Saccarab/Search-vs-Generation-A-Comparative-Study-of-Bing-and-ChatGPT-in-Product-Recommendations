from openpyxl import load_workbook

def mark_sidepanel_only(xlsx_path):
    wb = load_workbook(xlsx_path)
    ws = wb['urls']
    headers = [cell.value for cell in ws[1]]
    
    if 'missing_reason' not in headers:
        print("Column 'missing_reason' not found.")
        return
        
    mr_idx = headers.index('missing_reason') + 1
    
    count = 0
    for r in range(2, ws.max_row + 1):
        reason = str(ws.cell(row=r, column=mr_idx).value or "")
        if "Additional-only" in reason:
            ws.cell(row=r, column=mr_idx).value = "Side-Panel Only (Skipped)"
            count += 1
            
    wb.save(xlsx_path)
    print(f"Updated {count} rows to 'Side-Panel Only (Skipped)'.")

if __name__ == "__main__":
    mark_sidepanel_only('geo-fresh.xlsx')
