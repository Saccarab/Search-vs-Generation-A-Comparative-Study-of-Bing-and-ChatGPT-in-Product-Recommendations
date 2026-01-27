import argparse
from pathlib import Path
from urllib.parse import urlsplit

from openpyxl import load_workbook


def normalize_domain(raw):
    if not isinstance(raw, str) or not raw.strip():
        return ""
    raw = raw.strip()
    if "://" not in raw:
        raw = "https://" + raw
    try:
        host = urlsplit(raw).netloc.lower()
        if host.startswith("www."):
            host = host[4:]
        for sub in ["apps.", "play.", "chrome.", "itunes.", "microsoft.", "addons."]:
            if host.startswith(sub):
                host = host[len(sub):]
        return host
    except Exception:
        return ""


def get_cell(row, idx):
    if idx is None:
        return ""
    value = row[idx].value
    return value if value is not None else ""


def main():
    parser = argparse.ArgumentParser(description="Fix listicle_products.is_host_domain based on domains.")
    parser.add_argument("--xlsx", default="geo-fresh.xlsx", help="Input workbook path")
    parser.add_argument("--out", default="", help="Output workbook path (default: *_is_host_domain_fixed.xlsx)")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise SystemExit(f"Workbook not found: {xlsx_path}")

    out_path = Path(args.out) if args.out else xlsx_path.with_name(
        f"{xlsx_path.stem}_is_host_domain_fixed.xlsx"
    )

    wb = load_workbook(xlsx_path)
    if "listicle_products" not in wb.sheetnames:
        raise SystemExit("Workbook missing sheet: listicle_products")

    ws = wb["listicle_products"]
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    header_map = {}
    for idx, cell in enumerate(header_row):
        key = (cell.value or "").strip().lower()
        if key:
            header_map[key] = idx

    def col_idx(name):
        return header_map.get(name)

    if "is_host_domain" not in header_map:
        new_col_idx = ws.max_column
        ws.cell(row=1, column=new_col_idx + 1, value="is_host_domain")
        header_map["is_host_domain"] = new_col_idx

    idx_listicle_url = col_idx("listicle_url")
    idx_listicle_domain = col_idx("listicle_domain")
    idx_product_url = col_idx("product_url")
    idx_product_domain = col_idx("product_domain")
    idx_is_host = col_idx("is_host_domain")

    updated = 0
    total = 0
    host_true = 0

    for row in ws.iter_rows(min_row=2):
        total += 1
        listicle_domain = normalize_domain(get_cell(row, idx_listicle_domain))
        if not listicle_domain:
            listicle_domain = normalize_domain(get_cell(row, idx_listicle_url))
        product_domain = normalize_domain(get_cell(row, idx_product_domain))
        if not product_domain:
            product_domain = normalize_domain(get_cell(row, idx_product_url))

        is_host = 1 if listicle_domain and product_domain and listicle_domain == product_domain else 0
        if is_host:
            host_true += 1

        row[idx_is_host].value = is_host
        updated += 1

    wb.save(out_path)
    print(f"Updated {updated} rows (host_domain=1 for {host_true}).")
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
