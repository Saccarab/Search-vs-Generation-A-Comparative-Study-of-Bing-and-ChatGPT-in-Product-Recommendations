import openpyxl
import pandas as pd
from pathlib import Path

def cleanup_failed_runs(xlsx_path, chatgpt_csv_path):
    # 1. Identify Run IDs to remove
    # Dynamic detection: any run in the Excel 'runs' sheet with missing rewritten_query
    FAILED_RUN_IDS = []
    
    if Path(xlsx_path).exists():
        df_runs = pd.read_excel(xlsx_path, sheet_name='runs')
        # Also include runs where rewritten_query is the literal string "nan" or "N/A"
        broken = df_runs[
            df_runs['rewritten_query'].isna() | 
            (df_runs['rewritten_query'].astype(str).str.lower() == 'nan') |
            (df_runs['rewritten_query'].astype(str).str.upper() == 'N/A')
        ]
        FAILED_RUN_IDS = broken['run_id'].unique().tolist()
    
    # Add the historically known broken ones just in case
    HISTORICAL_FAILED = [
        "P002_r2", "P021_r1", "P021_r3", "P026_r1", "P031_r3", 
        "P034_r2", "P035_r3", "P046_r3", "P071_r3", "P076_r1",
        "P009_r2", "P011_r2", "P029_r2", "P048_r3", "P060_r3", 
        "P062_r1", "P062_r2", "P062_r3"
    ]
    for rid in HISTORICAL_FAILED:
        if rid not in FAILED_RUN_IDS:
            FAILED_RUN_IDS.append(rid)

    if not FAILED_RUN_IDS:
        print("No failed runs detected.")
        return

    print(f"Starting cleanup for {len(FAILED_RUN_IDS)} identified runs: {FAILED_RUN_IDS}")

    # --- Clean Excel ---
    if Path(xlsx_path).exists():
        wb = openpyxl.load_workbook(xlsx_path)
        
        # Clean runs, citations, bing_results
        for sheet_name in ["runs", "citations", "bing_results"]:
            if sheet_name not in wb.sheetnames: continue
            ws = wb[sheet_name]
            header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            if "run_id" not in header: continue
            rid_idx = header.index("run_id") + 1
            
            rows_to_delete = []
            for r in range(2, ws.max_row + 1):
                rid = str(ws.cell(r, rid_idx).value or "").strip()
                if rid in FAILED_RUN_IDS:
                    rows_to_delete.append(r)
            
            for r in reversed(rows_to_delete):
                ws.delete_rows(r)
            print(f"  - Excel [{sheet_name}]: Deleted {len(rows_to_delete)} rows.")

        # Clean URLs sheet (remove URLs that are no longer used anywhere)
        if "urls" in wb.sheetnames:
            ws_urls = wb["urls"]
            # 1. Find all URLs currently in use
            used_urls = set()
            for sn in ["citations", "bing_results"]:
                ws = wb[sn]
                h = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
                if "url" not in h: continue
                u_idx = h.index("url") + 1
                for r in range(2, ws.max_row + 1):
                    val = ws.cell(r, u_idx).value
                    if val: used_urls.add(str(val).strip())
            
            # 2. Delete unused from urls sheet
            h_urls = [ws_urls.cell(1, c).value for c in range(1, ws_urls.max_column + 1)]
            url_idx = h_urls.index("url") + 1
            rows_to_delete = []
            for r in range(2, ws_urls.max_row + 1):
                u = str(ws_urls.cell(r, url_idx).value or "").strip()
                if u and u not in used_urls:
                    rows_to_delete.append(r)
            
            for r in reversed(rows_to_delete):
                ws.delete_rows(r)
            print(f"  - Excel [urls]: Deleted {len(rows_to_delete)} orphaned URL rows.")

        wb.save(xlsx_path)
        print(f"Saved {xlsx_path}")

    # --- Clean ChatGPT CSV (Optional, only if path provided) ---
    if chatgpt_csv_path and Path(chatgpt_csv_path).exists():
        try:
            df = pd.read_csv(chatgpt_csv_path, engine="python")
            df['temp_run_id'] = df['prompt_id'].astype(str) + "_r" + df['run_number'].astype(str)
            
            before_len = len(df)
            df_clean = df[~df['temp_run_id'].isin(FAILED_RUN_IDS)].copy()
            df_clean.drop(columns=['temp_run_id'], inplace=True)
            
            if len(df_clean) < before_len:
                df_clean.to_csv(chatgpt_csv_path, index=False)
                print(f"  - CSV [{chatgpt_csv_path}]: Removed {before_len - len(df_clean)} rows.")
        except Exception as e:
            print(f"  - CSV cleanup failed: {e}")

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--chatgpt-csv", required=False)
    args = parser.parse_args()
    cleanup_failed_runs(args.xlsx, args.chatgpt_csv)
