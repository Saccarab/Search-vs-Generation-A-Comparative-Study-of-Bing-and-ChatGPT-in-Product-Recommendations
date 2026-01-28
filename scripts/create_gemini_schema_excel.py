#!/usr/bin/env python3
"""
Create initial Gemini grounding analysis Excel file with proper schema.

This creates gemini_grounding_analysis.xlsx with the same structure as geo_updated.xlsx
but adapted for Gemini AI Search Filter research.
"""

import argparse
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill
import datetime as dt

def create_sheet_with_schema(workbook, sheet_name, columns, descriptions):
    """Create a worksheet with proper column headers and documentation."""
    ws = workbook.create_sheet(sheet_name)

    # Add column headers
    for col_num, column_name in enumerate(columns.keys(), 1):
        cell = ws.cell(row=1, column=col_num, value=column_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFE6E6FA", end_color="FFE6E6FA", fill_type="solid")

    # Add descriptions as comments or in row 2
    for col_num, description in enumerate(descriptions.values(), 1):
        ws.cell(row=2, column=col_num, value=description)

    # Set column widths
    for col_num, column_name in enumerate(columns.keys(), 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = max(15, len(column_name) + 2)

    return ws

def create_gemini_analysis_excel(output_path):
    """Create the master Gemini analysis Excel file."""

    workbook = openpyxl.Workbook()

    # Remove default sheet
    workbook.remove(workbook.active)

    # Sheet 1: gemini_runs
    runs_columns = {
        'run_id': 'Unique run identifier (e.g., G0001_r1)',
        'query_id': 'Query identifier',
        'run_number': 'Run number (1, 2, 3) for consistency checks',
        'original_prompt': 'Original user query/prompt',
        'web_search_triggered': 'Whether Gemini triggered web search (boolean)',
        'response_timestamp': 'When Gemini API call was made (ISO format)',
        'model_version': 'Gemini model used (e.g., gemini-2.0-flash-exp)',
        'total_web_search_queries': 'Number of search queries Gemini generated',
        'total_grounding_chunks': 'Number of sources retrieved',
        'total_grounding_supports': 'Number of sources actually cited'
    }
    create_sheet_with_schema(workbook, 'gemini_runs', runs_columns, runs_columns)

    # Sheet 2: gemini_web_search_queries
    search_queries_columns = {
        'run_id': 'Foreign key to gemini_runs',
        'query_index': 'Position in Geminis query list (0-indexed)',
        'search_query': 'The actual search query Gemini sent to Google',
        'serp_results_count': 'Number of Google SERP results scraped for this query'
    }
    create_sheet_with_schema(workbook, 'gemini_web_search_queries', search_queries_columns, search_queries_columns)

    # Sheet 3: gemini_grounding_chunks (retrieved sources)
    chunks_columns = {
        'run_id': 'Foreign key to gemini_runs',
        'chunk_index': 'Position in groundingChunks array',
        'title': 'Source page title',
        'uri': 'Source URL',
        'domain': 'Extracted domain',
        'chunk_text': 'The actual text chunk Gemini extracted',
        'chunk_word_count': 'Word count of extracted chunk',
        'is_cited': 'Boolean: does this chunk appear in groundingSupports?'
    }
    create_sheet_with_schema(workbook, 'gemini_grounding_chunks', chunks_columns, chunks_columns)

    # Sheet 4: gemini_grounding_supports (cited sources)
    supports_columns = {
        'run_id': 'Foreign key to gemini_runs',
        'support_index': 'Position in groundingSupports array',
        'segment_text': 'The sentence/segment in Geminis response',
        'segment_word_count': 'Word count of the segment',
        'confidence_score': 'Geminis confidence score (0.0-1.0)',
        'grounding_chunk_indices': 'Array of chunk indices this segment cites',
        'citation_count': 'Number of chunks cited by this segment',
        'uri': 'Primary cited URL (from groundingChunkIndices[0])',
        'domain': 'Extracted domain of primary citation'
    }
    create_sheet_with_schema(workbook, 'gemini_grounding_supports', supports_columns, supports_columns)

    # Sheet 5: google_serp_results (control group)
    serp_columns = {
        'serp_query': 'The Gemini webSearchQuery this SERP is for',
        'run_id': 'Foreign key to gemini_runs',
        'position': 'SERP position (1-20)',
        'title': 'Result title',
        'url': 'Result URL',
        'display_url': 'Display URL shown on SERP',
        'snippet': 'Result snippet',
        'domain': 'Extracted domain',
        'serp_timestamp': 'When SERP was scraped (ISO format)'
    }
    create_sheet_with_schema(workbook, 'google_serp_results', serp_columns, serp_columns)

    # Sheet 6: gemini_citations (derived analysis)
    citations_columns = {
        'run_id': 'Foreign key to gemini_runs',
        'citation_type': 'grounding_support (Gemini equivalent of citations)',
        'url': 'Cited URL',
        'serp_rank': 'Position in Google SERP (null if not in top 20)',
        'chunk_index': 'Which groundingChunk this citation came from',
        'support_index': 'Which groundingSupport cited this',
        'segment_text': 'The response segment that cited this',
        'citation_group_size': 'How many URLs cited in this segment',
        'citation_in_group_rank': 'Position within the citation group'
    }
    create_sheet_with_schema(workbook, 'gemini_citations', citations_columns, citations_columns)

    # Sheet 7: urls (reuse existing schema with Gemini additions)
    urls_columns = {
        'url': 'Normalized URL (primary key)',
        'domain': 'Extracted domain',
        'content_path': 'Path to .txt content file',
        'meta_path': 'Path to .meta.json file',
        'content_word_count': 'Word count of extracted content',
        'has_schema_markup': 'Boolean',
        'fetched_at': 'Timestamp of content fetch',
        'page_title': 'Extracted page title',
        'meta_description': 'Meta description',
        'canonical_url': 'Canonical URL',
        'published_date': 'Publication date',
        'modified_date': 'Last modified date',
        'type': 'Gemini label (listicle, product, blog, etc.)',
        'content_format': 'Gemini label',
        'tone': 'Gemini label',
        'promotional_intensity_score': '0-10 score',
        'gemini_chunk_extracted': 'Boolean: was this URL extracted by Gemini?',
        'gemini_support_cited': 'Boolean: was this URL cited in Gemini response?',
        'gemini_chunk_word_count': 'Words extracted by Gemini from this URL',
        'gemini_extraction_efficiency': '(gemini_chunk_word_count / content_word_count) * 100'
    }
    create_sheet_with_schema(workbook, 'urls', urls_columns, urls_columns)

    # Add metadata sheet
    metadata_ws = workbook.create_sheet('metadata')
    metadata_ws.cell(row=1, column=1, value='Gemini Grounding Analysis Workbook')
    metadata_ws.cell(row=2, column=1, value='Created:')
    metadata_ws.cell(row=2, column=2, value=dt.datetime.now().isoformat())
    metadata_ws.cell(row=3, column=1, value='Purpose:')
    metadata_ws.cell(row=3, column=2, value='AI Search Filter analysis inspired by Dejan AI research')
    metadata_ws.cell(row=4, column=1, value='Schema Version:')
    metadata_ws.cell(row=4, column=2, value='1.0')

    # Save the workbook
    workbook.save(output_path)
    print(f"Created Gemini analysis Excel file: {output_path}")

    # Print schema summary
    print(f"\nSchema Summary:")
    print(f"- {len(workbook.sheetnames)} sheets created")
    print(f"- Sheets: {', '.join(workbook.sheetnames)}")

def main():
    parser = argparse.ArgumentParser(description='Create Gemini grounding analysis Excel file')
    parser.add_argument('--output', default='data/gemini_grounding_analysis.xlsx',
                        help='Output Excel file path')

    args = parser.parse_args()

    # Ensure data directory exists
    Path(args.output).parent.mkdir(exist_ok=True)

    create_gemini_analysis_excel(args.output)

if __name__ == "__main__":
    main()