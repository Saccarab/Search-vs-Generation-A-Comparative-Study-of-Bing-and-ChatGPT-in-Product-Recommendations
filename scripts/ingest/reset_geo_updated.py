"""
Reset selected sheets in geo_updated.xlsx by clearing all rows below the header.

This is intended for templates that contain mock/sample data or preformatted empty rows
that make new ingests appear "below" the visible area.

Sheets reset (by default):
  - runs
  - citations
  - bing_results
  - urls

Header row (row 1) is preserved.
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import List

import openpyxl


DEFAULT_SHEETS = ["runs", "citations", "bing_results", "urls"]


def clear_sheet_below_header(ws) -> int:
    """Clear values for all cells below header; returns number of cleared rows (data rows)."""
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row <= 1:
        return 0
    for r in range(2, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(r, c).value = None
    # Keep max_row stable; Excel may still show formatting, but values are gone.
    return max_row - 1


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Path to geo_updated.xlsx")
    ap.add_argument(
        "--sheets",
        default=",".join(DEFAULT_SHEETS),
        help=f"Comma-separated sheet names to reset (default: {','.join(DEFAULT_SHEETS)})",
    )
    args = ap.parse_args()

    xlsx = Path(args.xlsx)
    wb = openpyxl.load_workbook(xlsx)
    sheets: List[str] = [s.strip() for s in str(args.sheets).split(",") if s.strip()]

    for name in sheets:
        if name not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {name}")

    stats = {}
    for name in sheets:
        ws = wb[name]
        stats[name] = clear_sheet_below_header(ws)

    wb.save(xlsx)
    print(f"Reset {xlsx}")
    for name, n in stats.items():
        print(f"- {name}: cleared {n} data rows")


if __name__ == "__main__":
    main()



