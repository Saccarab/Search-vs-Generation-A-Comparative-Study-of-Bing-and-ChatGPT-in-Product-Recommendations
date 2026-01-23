import openpyxl
from pathlib import Path

def patch_missing_queries(xlsx_path, query_map):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["runs"]
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    
    rid_idx = header.index("run_id") + 1
    query_idx = header.index("rewritten_query") + 1
    
    patched_count = 0
    for r in range(2, ws.max_row + 1):
        rid = str(ws.cell(r, rid_idx).value or "").strip()
        if rid in query_map:
            ws.cell(r, query_idx).value = query_map[rid]
            patched_count += 1
            print(f"Patched {rid} -> {query_map[rid]}")
            
    wb.save(xlsx_path)
    print(f"Successfully patched {patched_count} runs in {xlsx_path}")

if __name__ == "__main__":
    query_map = {
        "P002_r2": "text to speech services that create conversations with multiple voices",
        "P026_r1": "video translator that can transcribe Spanish videos into English software",
        "P034_r2": "Windows app live translation during calls Google Meet or other meetings live translation software",
        "P076_r1": "best free AI video translator tools",
        "P021_r1": "free website or program to translate a 20 minute video and add subtitles automatically",
        "P021_r3": "free website program translate video add subtitles 20 minute video translate and subtitle free",
        "P031_r3": "free audio to text transcription services available currently",
        "P035_r3": "translation services with live interpreters 2-day pricing live interpreter translation services pricing",
        "P046_r3": "Doctranslate competitors similar translation services text audio video live translation",
        "P071_r3": "best transcription websites to convert audio to text"
    }
    patch_missing_queries("geo-fresh.xlsx", query_map)
