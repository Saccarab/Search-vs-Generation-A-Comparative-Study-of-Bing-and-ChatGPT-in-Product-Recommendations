"""
Validate a data collection / ingest iteration for geo_updated.xlsx.

Checks:
- bing_results has rows + page_num present
- citations has rows (if present)
- urls has non-empty url + dedupe sanity (normalized-key duplicates)
- ck/a redirect URLs removed
- content_path files exist (best-effort)
"""

from __future__ import annotations

import argparse
import os
from pathlib import Path
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

import openpyxl


TRACKING_PARAM_PREFIXES = ("utm_",)
TRACKING_PARAMS_EXACT = {
    "gclid",
    "fbclid",
    "msclkid",
    "yclid",
    "mc_cid",
    "mc_eid",
    "igshid",
    "ref",
    "ref_src",
    "spm",
}


def normalize_url_key(raw_url: str) -> str:
    if not raw_url:
        return ""
    u = str(raw_url).strip()
    if not u:
        return ""
    if "://" not in u:
        u = "https://" + u
    p = urlsplit(u)
    host = (p.netloc or "").lower()
    if host.startswith("www."):
        host = host[4:]
    path = p.path or "/"
    if len(path) > 1 and path.endswith("/"):
        path = path[:-1]
    kept = []
    for k, v in parse_qsl(p.query, keep_blank_values=True):
        lk = k.lower()
        if any(lk.startswith(px) for px in TRACKING_PARAM_PREFIXES):
            continue
        if lk in TRACKING_PARAMS_EXACT:
            continue
        kept.append((k, v))
    q = urlencode(kept, doseq=True)
    return urlunsplit(("", host, path, q, "")).lstrip("/")


def header_map(ws):
    return {
        str(ws.cell(1, c).value).strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value is not None and str(ws.cell(1, c).value).strip()
    }


def count_nonempty(ws, col_idx: int) -> int:
    n = 0
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, col_idx).value
        if v is None:
            continue
        if str(v).strip() == "":
            continue
        n += 1
    return n


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="geo_updated.xlsx")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)

    print(f"Workbook: {args.xlsx}")

    # ---- bing_results ----
    ws_bing = wb["bing_results"]
    hb = header_map(ws_bing)
    bing_rows = count_nonempty(ws_bing, hb["run_id"]) if "run_id" in hb else 0
    page_num_nonempty = count_nonempty(ws_bing, hb["page_num"]) if "page_num" in hb else 0
    print("\n== bing_results ==")
    print("rows (by run_id):", bing_rows)
    print("page_num nonempty:", page_num_nonempty)

    # ck/a in bing_results.url
    ck_bing = 0
    if "url" in hb:
        for r in range(2, ws_bing.max_row + 1):
            v = ws_bing.cell(r, hb["url"]).value
            if v and "bing.com/ck/a" in str(v).lower():
                ck_bing += 1
    print("ck/a remaining:", ck_bing)

    # ---- citations ----
    print("\n== citations ==")
    if "citations" in wb.sheetnames:
        ws_cit = wb["citations"]
        hc = header_map(ws_cit)
        cit_rows = count_nonempty(ws_cit, hc["run_id"]) if "run_id" in hc else 0
        print("rows (by run_id):", cit_rows)
    else:
        print("sheet missing (ok if you haven't ingested ChatGPT yet)")

    # ---- urls ----
    ws_urls = wb["urls"]
    hu = header_map(ws_urls)
    urls_nonempty = count_nonempty(ws_urls, hu["url"]) if "url" in hu else 0
    cp_nonempty = count_nonempty(ws_urls, hu["content_path"]) if "content_path" in hu else 0
    print("\n== urls ==")
    print("urls nonempty:", urls_nonempty)
    print("content_path nonempty:", cp_nonempty)

    # ck/a + duplicates
    ck_urls = 0
    keys_seen = set()
    dup_keys = 0
    for r in range(2, ws_urls.max_row + 1):
        u = ws_urls.cell(r, hu["url"]).value
        if not u:
            continue
        s = str(u).strip()
        if not s:
            continue
        if "bing.com/ck/a" in s.lower():
            ck_urls += 1
        k = normalize_url_key(s)
        if k in keys_seen:
            dup_keys += 1
        else:
            keys_seen.add(k)
    print("ck/a remaining:", ck_urls)
    print("duplicate normalized-key rows:", dup_keys)

    # content_path existence check (sampled)
    missing_files = 0
    checked = 0
    if "content_path" in hu:
        for r in range(2, ws_urls.max_row + 1):
            cp = ws_urls.cell(r, hu["content_path"]).value
            if not cp:
                continue
            p = str(cp).strip()
            if not p:
                continue
            checked += 1
            if not os.path.exists(p):
                missing_files += 1
            if checked >= 50:  # sample first 50 to keep it fast
                break
    print("content_path file existence (sample 50):", f"{checked-missing_files}/{checked} exist")

    # PASS/FAIL summary
    print("\n== summary ==")
    ok = True
    if bing_rows == 0:
        ok = False
        print("FAIL: bing_results has 0 rows")
    if ck_bing != 0 or ck_urls != 0:
        ok = False
        print("FAIL: ck/a URLs still present")
    if dup_keys != 0:
        ok = False
        print("FAIL: urls has duplicate normalized keys")
    if ok:
        print("PASS: iteration looks consistent (SERP ingested, redirects cleaned, urls deduped).")
    return 0 if ok else 2


if __name__ == "__main__":
    raise SystemExit(main())

