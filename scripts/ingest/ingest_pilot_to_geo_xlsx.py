"""
Ingest pilot datasets into geo_updated.xlsx (raw + enrichment-ready, no inference).

Inputs:
  - ChatGPT results CSV (ChatGPT scraper export)
  - Bing results CSV (cleaned top<=30 export; deduped, sequential ranks)
  - Target Excel workbook (geo_updated.xlsx)

Writes:
  - Appends/updates:
      - runs
      - citations (from sources_cited_json and sources_additional_json)
      - bing_results
      - urls (unique URLs seen in citations + bing_results)

Notes:
  - We keep enrichment fields in `urls` blank (wordcount/tone/etc.) for later pipelines.
  - We do NOT write overlap metrics to Excel (analysis lives in Python/BigQuery).
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import base64
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from urllib.parse import urlsplit

import openpyxl
import pandas as pd


def utc_now_iso() -> str:
    return dt.datetime.utcnow().replace(microsecond=0).isoformat() + "Z"

def safe_str(v: object) -> str:
    """Convert pandas/Excel values to clean strings (avoid literal 'nan')."""
    if v is None:
        return ""
    try:
        # pandas uses NaN float for missing values
        if pd.isna(v):
            return ""
    except Exception:
        pass
    s = str(v).strip()
    if s.lower() == "nan":
        return ""
    return s

def decode_bing_ck_url(raw_url: str) -> str:
    """
    Decode Bing click-tracking redirect URLs (bing.com/ck/a) to their real destination if possible.
    If we can't decode to a real http(s) URL, return empty string.
    """
    if not raw_url or not isinstance(raw_url, str):
        return ""
    url = raw_url.strip()
    if not url:
        return ""
    try:
        parts = urlsplit(url if "://" in url else "https://" + url)
        host = (parts.netloc or "").lower()
        if host.startswith("www."):
            host = host[4:]
        if not host.endswith("bing.com") or not parts.path.startswith("/ck/a"):
            return url
        # parse query params
        qs = dict(pd.compat.urlsafe_parse_qsl(parts.query)) if hasattr(pd.compat, "urlsafe_parse_qsl") else dict()
        if not qs:
            # fallback parse without pandas compat
            from urllib.parse import parse_qsl
            qs = dict(parse_qsl(parts.query, keep_blank_values=True))
        u = (qs.get("u") or "").strip()
        if not u.startswith("a1"):
            return ""
        b64 = u[2:]
        pad = "=" * ((4 - (len(b64) % 4)) % 4)
        try:
            decoded = base64.urlsafe_b64decode((b64 + pad).encode("utf-8")).decode("utf-8", errors="replace")
        except Exception:
            return ""
        return decoded if decoded.startswith("http") else ""
    except Exception:
        return ""


def extract_domain(raw_url: str) -> str:
    if not raw_url or not isinstance(raw_url, str):
        return ""
    url = raw_url.strip()
    if not url:
        return ""
    if "://" not in url:
        url = "https://" + url
    host = (urlsplit(url).netloc or "").lower()
    if host.startswith("www."):
        host = host[4:]
    return host


def parse_sources_json(cell: object) -> List[Dict[str, str]]:
    """Parse sources_*_json cell into list of dicts with url/title/description/domain (best-effort)."""
    if cell is None:
        return []
    s = str(cell).strip()
    if not s or s == "[]":
        return []
    try:
        arr = json.loads(s)
    except Exception:
        return []
    out: List[Dict[str, str]] = []
    if isinstance(arr, list):
        for it in arr:
            if isinstance(it, dict) and isinstance(it.get("url"), str):
                url = it["url"].strip()
                out.append(
                    {
                        "url": url,
                        "title": str(it.get("title") or "").strip(),
                        "description": str(it.get("description") or "").strip(),
                        "domain": str(it.get("domain") or extract_domain(url)).strip(),
                    }
                )
            elif isinstance(it, str):
                url = it.strip()
                out.append({"url": url, "title": "", "description": "", "domain": extract_domain(url)})
    return out

def parse_items_json(cell: object) -> List[dict]:
    if cell is None:
        return []
    s = str(cell).strip()
    if not s or s == "[]":
        return []
    try:
        arr = json.loads(s)
    except Exception:
        return []
    return arr if isinstance(arr, list) else []


def sheet_header_map(ws) -> Dict[str, int]:
    """Map header name -> 1-based column index using first row."""
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    m = {}
    for i, v in enumerate(header, start=1):
        if v is None:
            continue
        key = str(v).strip()
        if key:
            m[key] = i
    return m


def find_row_by_key(ws, col_idx: int, key: str) -> Optional[int]:
    """Return row number where ws[row, col_idx] == key, else None."""
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=col_idx).value
        if v is None:
            continue
        if str(v).strip() == key:
            return r
    return None


def append_row(ws, header: Dict[str, int], row: Dict[str, object]) -> int:
    # Many Excel templates have formatting pre-applied down to row 1000+,
    # so ws.max_row can point to a "sample area" even if cells are empty.
    # We want the first truly empty row (based on whether any header column has a value).
    r = None
    last = ws.max_row
    for rr in range(2, last + 1):
        has_any = False
        for col_idx in header.values():
            v = ws.cell(row=rr, column=col_idx).value
            if v is not None and str(v).strip() != "":
                has_any = True
                break
        if not has_any:
            r = rr
            break
    if r is None:
        r = ws.max_row + 1
    for k, v in row.items():
        if k not in header:
            continue
        ws.cell(row=r, column=header[k]).value = v
    return r


def upsert_row(ws, header: Dict[str, int], key_col: str, key_val: str, row: Dict[str, object]) -> int:
    key_idx = header[key_col]
    existing = find_row_by_key(ws, key_idx, key_val)
    if existing is None:
        row = dict(row)
        row[key_col] = key_val
        return append_row(ws, header, row)
    # update only non-empty values
    for k, v in row.items():
        if k not in header:
            continue
        if v is None or (isinstance(v, str) and v.strip() == ""):
            continue
        ws.cell(row=existing, column=header[k]).value = v
    return existing


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Path to geo_updated.xlsx")
    ap.add_argument("--chatgpt", required=True, help="Path to ChatGPT results CSV")
    ap.add_argument("--bing", required=True, help="Path to cleaned Bing results CSV (top<=30)")
    ap.add_argument("--user-location", default="", help="e.g., TR/Istanbul or DE/Berlin")
    ap.add_argument("--vpn-status", default="off", help="off|on|unknown")
    ap.add_argument("--ui-session", default="", help="Optional UI session label")
    ap.add_argument("--notes", default="", help="Optional notes written to runs.notes")
    args = ap.parse_args()

    xlsx = Path(args.xlsx)
    chat = pd.read_csv(Path(args.chatgpt), engine="python")
    bing = pd.read_csv(Path(args.bing), engine="python")

    wb = openpyxl.load_workbook(xlsx)
    ws_runs = wb["runs"]
    ws_cit = wb["citations"]
    ws_bing = wb["bing_results"]
    ws_urls = wb["urls"]

    h_runs = sheet_header_map(ws_runs)
    h_cit = sheet_header_map(ws_cit)
    h_bing = sheet_header_map(ws_bing)
    h_urls = sheet_header_map(ws_urls)

    # ---- runs ----
    chat["run_id"] = chat.apply(lambda r: f"{r['prompt_id']}_r{int(r['run_number'])}", axis=1)
    for _, r in chat.iterrows():
        run_id = str(r["run_id"])
        prompt_id = str(r["prompt_id"])
        rewritten = str(r.get("generated_search_query", "") or "").strip()
        # citation_count: count of cited URLs (not additional)
        cited = parse_sources_json(r.get("sources_cited_json"))
        citation_count = len({c["url"] for c in cited if c.get("url")})

        upsert_row(
            ws_runs,
            h_runs,
            "run_id",
            run_id,
            {
                "prompt_id": prompt_id,
                "run_ts_utc": utc_now_iso(),
                "citation_count": citation_count,
                "ui_session": args.ui_session,
                "rewritten_query": rewritten,
                "user_location": args.user_location,
                "notes": args.notes,
                "vpn_status": args.vpn_status,
            },
        )

    # ---- citations ----
    # We'll write one row per cited/additional URL per run.
    # cite_position: stable order within each section (1..N)
    # citation_type: 'cited' | 'additional'
    #
    # Additionally: we write item-level inline citations from items_json so you can analyze
    # citation-to-item linkage (item_section_title/item_position/item_name/item_text + group size/rank).
    # These use:
    #   citation_type = 'inline'
    #   capture_method = 'inline_chips'
    #   citation_scope = 'inline'
    # Build existing citation keys so re-running is idempotent
    existing_citation_keys = set()
    if "run_id" in h_cit and "url" in h_cit and "citation_type" in h_cit:
        for rr in range(2, ws_cit.max_row + 1):
            rid = ws_cit.cell(rr, h_cit["run_id"]).value
            url = ws_cit.cell(rr, h_cit["url"]).value
            ctype = ws_cit.cell(rr, h_cit["citation_type"]).value
            if rid and url and ctype:
                key = (str(rid).strip(), str(url).strip(), str(ctype).strip(),
                       str(ws_cit.cell(rr, h_cit.get("cite_position", 0)).value or "").strip(),
                       str(ws_cit.cell(rr, h_cit.get("item_position", 0)).value or "").strip(),
                       str(ws_cit.cell(rr, h_cit.get("citation_in_group_rank", 0)).value or "").strip())
                existing_citation_keys.add(key)

    for _, r in chat.iterrows():
        run_id = str(r["run_id"])

        cited = parse_sources_json(r.get("sources_cited_json"))
        additional = parse_sources_json(r.get("sources_additional_json"))
        items = parse_items_json(r.get("items_json"))

        # Map URL -> cite_position in sources panel (cited first, then additional)
        # NOTE: This is not "inline order". It's the position in the Sources panel list (1..N).
        sources_panel_pos: Dict[str, int] = {}
        for i, it in enumerate(cited, start=1):
            u = (it.get("url") or "").strip()
            if u and u not in sources_panel_pos:
                sources_panel_pos[u] = i
        for i, it in enumerate(additional, start=1):
            u = (it.get("url") or "").strip()
            if u and u not in sources_panel_pos:
                sources_panel_pos[u] = i

        # Helper: map URL -> title/description/domain from sources panel JSON (best-effort)
        url_meta = {}
        for it in cited + additional:
            u = it.get("url", "").strip()
            if u:
                url_meta[u] = it

        def emit(section: str, arr: List[Dict[str, str]]):
            for i, it in enumerate(arr, start=1):
                url = it.get("url", "").strip()
                if not url:
                    continue
                key = (run_id, url, section, str(i), "", "")
                if key in existing_citation_keys:
                    continue
                append_row(
                    ws_cit,
                    h_cit,
                    {
                        "run_id": run_id,
                        "cite_position": i,
                        "citation_title": it.get("title", ""),
                        "url": url,
                        "citation_domain": it.get("domain", "") or extract_domain(url),
                        "citation_type": section,
                        "capture_method": "sources_panel",
                        "notes": it.get("description", ""),
                        "citation_scope": "sources_panel",
                        # item_* columns intentionally left blank for now (can be filled from items_json later)
                    },
                )

        emit("cited", cited)
        emit("additional", additional)

        # Item-level inline citations (from items_json)
        for item in items:
            try:
                item_section_title = safe_str(item.get("item_section_title", ""))
                item_position = item.get("item_position", None)
                item_name = safe_str(item.get("item_name", ""))
                item_text = safe_str(item.get("item_text", ""))
                chip_groups = item.get("chip_groups", []) or []
                for group in chip_groups:
                    links = group.get("links", []) if isinstance(group, dict) else []
                    if not isinstance(links, list) or not links:
                        continue
                    group_size = len(links)
                    for j, link in enumerate(links, start=1):
                        url = safe_str(link)
                        if not url:
                            continue
                        # prefer metadata from sources panel if same URL exists
                        meta = url_meta.get(url, {})
                        title = safe_str(meta.get("title", "")) if isinstance(meta, dict) else ""
                        domain = safe_str(meta.get("domain", "")) if isinstance(meta, dict) else ""
                        if not domain:
                            domain = extract_domain(url)

                        sp_pos = sources_panel_pos.get(url, None)
                        key = (run_id, url, "inline", "", str(item_position or ""), str(j))
                        if key in existing_citation_keys:
                            continue

                        append_row(
                            ws_cit,
                            h_cit,
                            {
                                "run_id": run_id,
                                # For inline rows, we store the Sources-panel position (if that URL appears there).
                                # Inline ordering itself is captured by item_position + citation_in_group_rank.
                                "cite_position": sp_pos,
                                "citation_title": title,
                                "url": url,
                                "citation_domain": domain,
                                "citation_type": "inline",
                                "capture_method": "inline_chips",
                                "notes": "",
                                "citation_scope": "inline",
                                "item_section_title": item_section_title,
                                "item_position": item_position,
                                "item_name": item_name,
                                "item_text": item_text,
                                "citation_group_size": group_size,
                                "citation_in_group_rank": j,
                                # Optional column (if present in the workbook): explicit duplicate of cite_position
                                # to make analysis intent clearer without overloading cite_position semantics.
                                "sources_panel_cite_position": sp_pos,
                            },
                        )
            except Exception:
                continue

    # ---- bing_results ----
    # Map cleaned Bing columns to sheet columns
    # bing_results header: run_id, bing_query, result_rank, result_title, url, result_domain, snippet, page_type, captured_at, page_num
    # cleaned CSV includes: run_id, query, position, title, snippet, domain, url, ... (and metadata)
    # Build existing keys for idempotent bing_results insert
    existing_bing_keys = set()
    if "run_id" in h_bing and "result_rank" in h_bing and "url" in h_bing:
        for rr in range(2, ws_bing.max_row + 1):
            rid = ws_bing.cell(rr, h_bing["run_id"]).value
            rank = ws_bing.cell(rr, h_bing["result_rank"]).value
            url = ws_bing.cell(rr, h_bing["url"]).value
            if rid and rank and url:
                existing_bing_keys.add((str(rid).strip(), str(rank).strip(), str(url).strip()))

    for _, r in bing.iterrows():
        rid = safe_str(r.get("run_id", ""))
        rank = int(r.get("position"))
        url = safe_str(r.get("url", ""))
        if (rid, str(rank), url) in existing_bing_keys:
            continue
        append_row(
            ws_bing,
            h_bing,
            {
                "run_id": rid,
                "bing_query": safe_str(r.get("query", "")),
                "result_rank": rank,
                "result_title": safe_str(r.get("title", "")),
                "url": url,
                "result_domain": safe_str(r.get("domain", "")),
                "snippet": safe_str(r.get("snippet", "")),
                "captured_at": utc_now_iso(),
                "page_num": r.get("page_num", None),
                # leave page_type/page_num blank unless you populate them elsewhere
            },
        )

    # ---- urls ----
    # Insert unique URLs seen (from citations + bing results). Keep enrichment columns blank.
    existing_urls = set()
    url_col_idx = h_urls.get("url")
    if url_col_idx:
        for rr in range(2, ws_urls.max_row + 1):
            v = ws_urls.cell(row=rr, column=url_col_idx).value
            if v:
                existing_urls.add(str(v).strip())

    new_urls: List[str] = []

    # from citations
    for _, r in chat.iterrows():
        for col in ("sources_cited_json", "sources_additional_json"):
            for it in parse_sources_json(r.get(col)):
                u = it.get("url", "").strip()
                if u and u not in existing_urls:
                    existing_urls.add(u)
                    new_urls.append(u)

    # from bing
    for u in bing["url"].astype(str).tolist():
        u = safe_str(u)
        # Skip Bing internal click-tracking URLs in the URL universe; they are not real pages.
        if "/ck/a" in u and "bing.com" in u:
            decoded = decode_bing_ck_url(u)
            u = decoded or ""
        if u and u not in existing_urls:
            existing_urls.add(u)
            new_urls.append(u)

    for u in new_urls:
        append_row(
            ws_urls,
            h_urls,
            {
                "url": u,
                "domain": extract_domain(u),
                # type/content_* left blank
            },
        )

    out_path = xlsx  # in-place update
    wb.save(out_path)
    print(f"Updated {out_path}")
    print(f"Inserted urls: {len(new_urls)}")


if __name__ == "__main__":
    main()


