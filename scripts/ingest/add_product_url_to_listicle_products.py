"""
Add a `product_url` column to geo_updated.xlsx -> listicle_products sheet (if missing).

Rationale:
- `listicle_url` is the parent listicle page URL (foreign key).
- `product_domain` is a hostname; `product_url` stores a per-item URL when explicitly present.

Usage:
  python scripts/ingest/add_product_url_to_listicle_products.py --xlsx geo_updated.xlsx
"""

from __future__ import annotations

import argparse

import openpyxl


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx)
    ws = wb["listicle_products"]

    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        headers.append("" if v is None else str(v).strip())

    if "product_url" in headers:
        print("product_url already exists")
        return

    ws.cell(1, ws.max_column + 1).value = "product_url"
    wb.save(args.xlsx)
    print("Added product_url")


if __name__ == "__main__":
    main()

