import pandas as pd
from openpyxl import load_workbook
import os

def audit_missing_content(xlsx_path):
    print(f"Auditing missing content in {xlsx_path}...")
    
    # Load data
    df_urls = pd.read_excel(xlsx_path, sheet_name='urls')
    df_citations = pd.read_excel(xlsx_path, sheet_name='citations')
    df_bing = pd.read_excel(xlsx_path, sheet_name='bing_results')

    # Normalize URLs for matching
    df_urls['url_clean'] = df_urls['url'].astype(str).str.strip()
    df_citations['url_clean'] = df_citations['url'].astype(str).str.strip()
    df_bing['url_clean'] = df_bing['url'].astype(str).str.strip()

    # Categorize URLs by where they appear
    bing_urls = set(df_bing['url_clean'].unique())
    
    # Citation types
    inline_cited_urls = set(df_citations[df_citations['citation_type'].isin(['inline', 'cited'])]['url_clean'].unique())
    additional_urls = set(df_citations[df_citations['citation_type'] == 'additional']['url_clean'].unique())
    
    def get_reason(row):
        url = row['url_clean']
        existing_reason = str(row.get('missing_reason', "") or "").strip()
        if existing_reason == "Skipped (Manual unreachable)":
            return existing_reason
        has_path = pd.notna(row['content_path']) and str(row['content_path']).strip() != ""
        
        # Priority check: where does this URL live?
        is_inline = url in inline_cited_urls
        is_bing = url in bing_urls
        is_additional = url in additional_urls
        
        # Word count
        word_count = row.get('content_word_count', 0)
        if pd.isna(word_count): word_count = 0
        
        # Case 1: We have content, but is it "useful"?
        if has_path:
            # If it's a high priority URL but has very low word count, mark as thin
            if (is_inline or is_bing) and word_count < 100:
                return "Failed: Thin Content (<100 words)"
            return "" # Good content

        # Case 2: We have NO path. Why?
        was_attempted = pd.notna(row['fetched_at']) and str(row['fetched_at']).strip() != ""
        
        if not was_attempted:
            if is_inline or is_bing:
                return "Pending Fetch (High Priority)"
            if is_additional:
                return "Side-Panel Only (Skipped)"
            return "Pending Fetch (Uncategorized)"
            
        # If attempted but no path was saved:
        if is_additional and not is_inline and not is_bing:
            return "Side-Panel Only (Skipped)"
            
        return "Failed: Fetch Error (403/429/Timeout)"

    df_urls['missing_reason'] = df_urls.apply(get_reason, axis=1)

    # Write back to Excel
    wb = load_workbook(xlsx_path)
    ws = wb['urls']
    
    # Find or create missing_reason column
    headers = [cell.value for cell in ws[1]]
    if 'missing_reason' in headers:
        col_idx = headers.index('missing_reason') + 1
    else:
        col_idx = len(headers) + 1
        ws.cell(row=1, column=col_idx).value = 'missing_reason'
    
    # Map back to Excel rows
    url_to_reason = dict(zip(df_urls['url'], df_urls['missing_reason']))
    
    url_col_idx = headers.index('url') + 1
    for r in range(2, ws.max_row + 1):
        url_val = ws.cell(row=r, column=url_col_idx).value
        if url_val in url_to_reason:
            ws.cell(row=r, column=col_idx).value = url_to_reason[url_val]

    wb.save(xlsx_path)
    print(f"Done. Summary of missing reasons:")
    print(df_urls['missing_reason'].value_counts())

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True)
    args = parser.parse_args()
    audit_missing_content(args.xlsx)
