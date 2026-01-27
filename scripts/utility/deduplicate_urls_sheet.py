import openpyxl
from pathlib import Path
from urllib.parse import urlsplit

def normalize_url(url: str) -> str:
    if not url: return ""
    url = url.strip().lower()
    if "://" not in url:
        url = "https://" + url
    try:
        parts = urlsplit(url)
        netloc = parts.netloc
        if netloc.startswith("www."):
            netloc = netloc[4:]
        path = parts.path.rstrip("/")
        # return netloc + path + ("?" + parts.query if parts.query else "")
        return netloc + path
    except:
        return url

def deduplicate_urls_sheet(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["urls"]
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    url_idx = header.index("url") + 1
    
    seen = {} # normalized -> original row index
    rows_to_delete = []
    
    for r in range(2, ws.max_row + 1):
        url = ws.cell(r, url_idx).value
        if not url: continue
        norm = normalize_url(str(url))
        if norm in seen:
            rows_to_delete.append(r)
        else:
            seen[norm] = r
            
    # Delete from bottom up
    for r in reversed(rows_to_delete):
        ws.delete_rows(r)
        
    wb.save(xlsx_path)
    print(f"Deduplicated urls sheet. Deleted {len(rows_to_delete)} rows.")

if __name__ == "__main__":
    deduplicate_urls_sheet("geo-fresh.xlsx")
