import pandas as pd
from openpyxl import load_workbook

def fix_headers(xlsx_path):
    wb = load_workbook(xlsx_path)
    if 'urls' not in wb.sheetnames:
        print("Sheet 'urls' not found.")
        return

    ws = wb['urls']
    headers = [cell.value for cell in ws[1]]
    print(f"Original headers: {headers}")

    # 1. Strip whitespace from headers
    # 2. Add readability_score if missing
    new_headers = []
    for h in headers:
        if h and isinstance(h, str):
            new_headers.append(h.strip())
        else:
            new_headers.append(h)

    if 'readability_score' not in new_headers:
        # Check if there's an 'Unnamed' column we can reuse or just append
        unnamed_idx = -1
        for i, h in enumerate(new_headers):
            if h and str(h).startswith('Unnamed:'):
                unnamed_idx = i
                break
        
        if unnamed_idx != -1:
            print(f"Replacing {new_headers[unnamed_idx]} with readability_score")
            new_headers[unnamed_idx] = 'readability_score'
        else:
            new_headers.append('readability_score')

    # Update the sheet
    for i, h in enumerate(new_headers):
        ws.cell(row=1, column=i+1).value = h
    
    wb.save(xlsx_path)
    print(f"Updated headers: {new_headers}")

if __name__ == "__main__":
    fix_headers('geo-fresh.xlsx')
