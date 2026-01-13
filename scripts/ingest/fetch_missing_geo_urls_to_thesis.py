"""
Fetch missing URL content for rows in geo_updated.xlsx -> `urls` sheet.

This is the Python-first replacement for extension-based content extraction and
also covers citation-only URLs.

Behavior:
- Select URLs where `content_path` is blank (or overwrite enabled)
- Fetch HTML via requests (follow redirects)
- Extract plain text (simple HTML tag stripping)
- Write text file to `--content-root/<run_label>/<hash>.txt`
- Update `urls`:
    content_path, content_word_count, has_schema_markup, fetched_at, domain

Usage (Windows):
  python scripts/ingest/fetch_missing_geo_urls_to_thesis.py ^
    --xlsx geo_updated.xlsx ^
    --content-root "C:\\Users\\User\\Documents\\thesis\\python_content" ^
    --max-urls 200 ^
    --sleep 0.5
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import os
import re
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Optional
from urllib.parse import parse_qsl, urlencode, urlparse, urlsplit, urlunsplit

import openpyxl
import requests

try:
    # Optional but strongly recommended for consistent, extension-like extraction.
    # Install: pip install beautifulsoup4 lxml
    from bs4 import BeautifulSoup  # type: ignore
except Exception:  # pragma: no cover
    BeautifulSoup = None  # type: ignore


TRACKING_PARAM_PREFIXES = ("utm_",)
TRACKING_PARAMS_EXACT = {
    "gclid",
    "fbclid",
    "msclkid",
    "yclid",
    "mc_cid",
    "mc_eid",
    "igshid",
    "ref",
    "ref_src",
    "spm",
}

RE_SCRIPT_STYLE = re.compile(r"(?is)<(script|style|noscript)[^>]*>.*?</\\1>")
RE_TAGS = re.compile(r"<[^>]+>")
RE_SPACE = re.compile(r"[ \\t\\f\\v]+")
RE_NEWLINES = re.compile(r"(\\r\\n|\\r|\\n)+")
WORD_RE = re.compile(r"[A-Za-z0-9]+(?:'[A-Za-z0-9]+)?")


@dataclass
class FetchResult:
    html: str
    final_url: str
    status: int
    error: str


def utc_now_iso() -> str:
    return dt.datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def short_hash(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()[:16]


def normalize_url_key(raw_url: str) -> str:
    if not raw_url:
        return ""
    u = str(raw_url).strip()
    if not u:
        return ""
    if "://" not in u:
        u = "https://" + u
    p = urlsplit(u)
    host = (p.netloc or "").lower()
    if host.startswith("www."):
        host = host[4:]
    path = p.path or "/"
    if len(path) > 1 and path.endswith("/"):
        path = path[:-1]
    kept = []
    for k, v in parse_qsl(p.query, keep_blank_values=True):
        lk = k.lower()
        if any(lk.startswith(px) for px in TRACKING_PARAM_PREFIXES):
            continue
        if lk in TRACKING_PARAMS_EXACT:
            continue
        kept.append((k, v))
    q = urlencode(kept, doseq=True)
    return urlunsplit(("", host, path, q, "")).lstrip("/")


def extract_domain(raw_url: str) -> str:
    if not raw_url:
        return ""
    u = str(raw_url).strip()
    if not u:
        return ""
    if "://" not in u:
        u = "https://" + u
    try:
        host = (urlparse(u).netloc or "").lower()
        return host[4:] if host.startswith("www.") else host
    except Exception:
        return ""


BLOCK_TAGS = (
    "p",
    "div",
    "li",
    "br",
    "tr",
    "td",
    "th",
    "h1",
    "h2",
    "h3",
    "h4",
    "h5",
    "h6",
    "section",
    "article",
    "header",
    "footer",
    "main",
    "aside",
)


def _looks_like_css_or_boilerplate(line: str) -> bool:
    s = (line or "").strip()
    if not s:
        return True
    low = s.lower()

    # Common CSS / build artifact patterns that leak through naive stripping
    if low.startswith("@layer ") or low.startswith("@media ") or low.startswith("@supports "):
        return True
    if low.startswith(":root") or low.startswith("--") or "css" in low and "{".encode() and "}" in low:
        # keep simple: :root / CSS vars are never useful for content
        return True
    if "var(--" in low or "--color-" in low or "--font-" in low:
        return True
    if low.startswith("function(") or "webpack" in low or "window.__" in low:
        return True

    # If line is mostly punctuation/code-like, drop it
    letters = sum(ch.isalpha() for ch in s)
    nonletters = len(s) - letters
    if letters <= 5 and len(s) > 40:
        return True
    if letters > 0 and (nonletters / max(1, letters)) > 2.5:
        return True

    return False


def strip_html_to_text(html: str) -> str:
    """
    Extract *visible-ish* page text.

    Preference: match the Bing extension behavior:
      - parse HTML into a DOM-like tree
      - remove script/style/noscript
      - remove common non-content containers (nav/header/footer/aside + ad/cookie/popup-ish classes)
      - take textContent and collapse whitespace to a single line

    Fallback: a regex-based stripper (less accurate) if BeautifulSoup isn't installed.
    """
    if not html:
        return ""

    if BeautifulSoup is not None:
        try:
            soup = BeautifulSoup(html, "lxml")

            # remove script/style/noscript (extension parity)
            for el in soup.select("script, style, noscript"):
                el.decompose()

            # remove common non-content tags (extension parity)
            for el in soup.select("nav, header, footer, aside"):
                el.decompose()

            # remove common non-content classes/ids (extension parity-ish)
            junk_tokens = (
                "navigation",
                "nav",
                "menu",
                "sidebar",
                "advertisement",
                "ad",
                "ads",
                "cookie",
                "popup",
                "modal",
                "overlay",
            )
            # If any token appears in class/id, drop the node
            for el in soup.find_all(True):
                try:
                    cid = " ".join(
                        filter(
                            None,
                            [
                                " ".join(el.get("class", []) or []),
                                str(el.get("id") or ""),
                            ],
                        )
                    ).lower()
                except Exception:
                    continue
                if not cid:
                    continue
                if any(tok in cid for tok in junk_tokens):
                    el.decompose()

            text = soup.get_text(" ", strip=True).replace("\u00a0", " ")
            text = re.sub(r"\\s+", " ", text).strip()
            return text
        except Exception:
            # fall through to regex fallback
            pass

    # ---- fallback (regex) ----
    html2 = RE_SCRIPT_STYLE.sub("\n", html)
    html2 = re.sub(r"(?is)<\\s*br\\s*/?>", "\n", html2)
    html2 = re.sub(r"(?is)</\\s*(%s)\\s*>" % "|".join(BLOCK_TAGS), "\n", html2)
    html2 = re.sub(r"(?is)<\\s*(%s)(\\s+[^>]*)?>" % "|".join(BLOCK_TAGS), "\n", html2)
    txt = RE_TAGS.sub(" ", html2)
    txt = txt.replace("\u00a0", " ")
    txt = RE_SPACE.sub(" ", txt)
    txt = RE_NEWLINES.sub("\n", txt)
    lines = [ln.strip() for ln in txt.split("\n")]
    lines = [ln for ln in lines if not _looks_like_css_or_boilerplate(ln)]
    return " ".join([ln for ln in lines if ln]).strip()


def count_words(text: str) -> int:
    if not text:
        return 0
    return len(WORD_RE.findall(text))


def has_schema_markup_from_html(html: str) -> bool:
    h = (html or "").lower()
    if 'type="application/ld+json"' in h:
        return True
    if "itemscope" in h and "itemtype" in h:
        return True
    return False


def sheet_header_map(ws) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        k = str(v).strip()
        if k:
            m[k] = c
    return m


def first_truly_empty_row(ws, header: Dict[str, int]) -> int:
    for rr in range(2, ws.max_row + 1):
        has_any = False
        for col_idx in header.values():
            v = ws.cell(rr, col_idx).value
            if v is not None and str(v).strip() != "":
                has_any = True
                break
        if not has_any:
            return rr
    return ws.max_row + 1


def fetch_html(url: str, timeout: int) -> FetchResult:
    try:
        headers = {"User-Agent": "Mozilla/5.0 (compatible; GEOThesisBot/0.1)"}
        resp = requests.get(url, headers=headers, timeout=timeout, allow_redirects=True)
        html = resp.text if resp.ok else ""
        return FetchResult(html=html, final_url=resp.url, status=resp.status_code, error="")
    except Exception as e:
        return FetchResult(html="", final_url=url, status=0, error=str(e))


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Path to geo_updated.xlsx")
    ap.add_argument("--content-root", required=True, help="Folder to write extracted text files into")
    ap.add_argument("--run-label", default="", help="Subfolder name under content-root (default: python_fetch_YYYYMMDD_HHMMSS)")
    ap.add_argument("--max-urls", type=int, default=0, help="If >0, process at most N URLs")
    ap.add_argument("--timeout", type=int, default=30, help="HTTP timeout seconds")
    ap.add_argument("--sleep", type=float, default=0.0, help="Sleep seconds between fetches")
    ap.add_argument("--overwrite", action="store_true", help="Overwrite existing content_path/fields")
    ap.add_argument("--min-text-chars", type=int, default=200, help="Minimum extracted text chars to accept as usable")
    ap.add_argument(
        "--only-url",
        action="append",
        default=[],
        help="Process only these exact URL(s). Can be passed multiple times. If set, overrides normal selection.",
    )
    ap.add_argument(
        "--url-contains",
        default="",
        help="Process only URLs whose raw URL string contains this substring (case-insensitive).",
    )
    args = ap.parse_args()

    xlsx = Path(args.xlsx)
    wb = openpyxl.load_workbook(xlsx)
    ws_urls = wb["urls"]
    hu = sheet_header_map(ws_urls)

    required = ["url", "domain", "content_path", "content_word_count", "has_schema_markup", "fetched_at"]
    missing = [c for c in required if c not in hu]
    if missing:
        raise SystemExit(f"urls sheet missing required columns: {missing}")

    label = args.run_label.strip()
    if not label:
        label = dt.datetime.now().strftime("python_fetch_%Y%m%d_%H%M%S")
    out_dir = Path(args.content_root) / label
    out_dir.mkdir(parents=True, exist_ok=True)

    processed = 0
    fetched_ok = 0
    wrote_files = 0
    skipped_has_content = 0
    failed = 0
    only_urls = {u.strip() for u in (args.only_url or []) if str(u).strip()}
    contains = (args.url_contains or "").strip().lower()

    for r in range(2, ws_urls.max_row + 1):
        raw_url = ws_urls.cell(r, hu["url"]).value
        if raw_url is None:
            continue
        url = str(raw_url).strip()
        if not url:
            continue

        if only_urls:
            if url not in only_urls:
                continue
        elif contains:
            if contains not in url.lower():
                continue

        existing_cp = str(ws_urls.cell(r, hu["content_path"]).value or "").strip()
        if existing_cp and not args.overwrite:
            skipped_has_content += 1
            continue

        processed += 1
        if args.max_urls and processed > args.max_urls:
            break

        fr = fetch_html(url, timeout=args.timeout)
        if fr.status != 200 or not fr.html:
            failed += 1
            # still stamp fetched_at so we know it was attempted
            ws_urls.cell(r, hu["fetched_at"]).value = utc_now_iso()
            if args.sleep:
                time.sleep(args.sleep)
            continue

        text = strip_html_to_text(fr.html)
        if len(text) < args.min_text_chars:
            failed += 1
            ws_urls.cell(r, hu["fetched_at"]).value = utc_now_iso()
            if args.sleep:
                time.sleep(args.sleep)
            continue

        key = normalize_url_key(url) or url
        fname = f"{short_hash(key)}.txt"
        fpath = out_dir / fname
        fpath.write_text(text, encoding="utf-8", errors="ignore")
        wrote_files += 1

        ws_urls.cell(r, hu["domain"]).value = ws_urls.cell(r, hu["domain"]).value or extract_domain(url)
        ws_urls.cell(r, hu["content_path"]).value = str(fpath)
        ws_urls.cell(r, hu["content_word_count"]).value = int(count_words(text))
        ws_urls.cell(r, hu["has_schema_markup"]).value = 1 if has_schema_markup_from_html(fr.html) else 0
        ws_urls.cell(r, hu["fetched_at"]).value = utc_now_iso()

        fetched_ok += 1
        if args.sleep:
            time.sleep(args.sleep)

        if processed % 25 == 0:
            print(f"processed {processed} (ok={fetched_ok}, failed={failed}, skipped={skipped_has_content})")

    wb.save(xlsx)
    print("Done.")
    print("output_dir:", out_dir)
    print("processed:", processed)
    print("fetched_ok:", fetched_ok)
    print("failed:", failed)
    print("skipped_has_content:", skipped_has_content)
    print("files_written:", wrote_files)


if __name__ == "__main__":
    main()

