"""
Clean literal 'nan' strings introduced by pandas when writing to Excel.

This is safe/idempotent: it only replaces cells whose text is exactly 'nan' (case-insensitive)
with blank (None).
"""

from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Path to geo_updated.xlsx")
    ap.add_argument("--sheet", default="bing_results", help="Sheet name (default: bing_results)")
    ap.add_argument("--column", default="snippet", help="Column header to clean (default: snippet)")
    args = ap.parse_args()

    xlsx = Path(args.xlsx)
    wb = openpyxl.load_workbook(xlsx)
    if args.sheet not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {args.sheet}")
    ws = wb[args.sheet]

    # header map
    header = {str(ws.cell(1, c).value).strip(): c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
    if args.column not in header:
        raise ValueError(f"Column not found: {args.column}. Available: {list(header.keys())}")
    col_idx = header[args.column]

    cleaned = 0
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, col_idx).value
        if isinstance(v, str) and v.strip().lower() == "nan":
            ws.cell(r, col_idx).value = None
            cleaned += 1

    wb.save(xlsx)
    print(f"Cleaned {cleaned} cells in {args.sheet}.{args.column} for {xlsx}")


if __name__ == "__main__":
    main()


