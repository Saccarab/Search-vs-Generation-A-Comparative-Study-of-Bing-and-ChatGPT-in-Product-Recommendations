"""
One-time cleanup for geo_updated.xlsx:
  - Decode Bing click-tracking URLs (bing.com/ck/a) in `urls.url` and `bing_results.url`
  - Collapse URL duplicates in `urls` by normalized key (http/https, fragments, tracking params, trailing slash)

This keeps the first row for each normalized key and merges non-empty fields into it,
then clears the duplicate rows (so formatting stays intact but data is not duplicated).
"""

from __future__ import annotations

import argparse
import base64
from typing import Dict, List, Optional, Tuple
from urllib.parse import parse_qsl, urlsplit, urlunsplit, urlencode

import openpyxl

TRACKING_PARAM_PREFIXES = ("utm_",)
TRACKING_PARAMS_EXACT = {
    "gclid",
    "fbclid",
    "msclkid",
    "yclid",
    "mc_cid",
    "mc_eid",
    "igshid",
    "ref",
    "ref_src",
    "spm",
}


def safe_str(v: object) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() == "nan" else s


def sheet_header_map(ws) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        k = str(v).strip()
        if k:
            m[k] = c
    return m


def normalize_url_key(raw_url: str) -> str:
    u = safe_str(raw_url)
    if not u:
        return ""
    if "://" not in u:
        u = "https://" + u
    p = urlsplit(u)
    host = (p.netloc or "").lower()
    if host.startswith("www."):
        host = host[4:]
    path = p.path or "/"
    if len(path) > 1 and path.endswith("/"):
        path = path[:-1]
    kept: List[Tuple[str, str]] = []
    for k, v in parse_qsl(p.query, keep_blank_values=True):
        lk = k.lower()
        if any(lk.startswith(px) for px in TRACKING_PARAM_PREFIXES):
            continue
        if lk in TRACKING_PARAMS_EXACT:
            continue
        kept.append((k, v))
    q = urlencode(kept, doseq=True)
    return urlunsplit(("", host, path, q, "")).lstrip("/")


def decode_bing_ck_url(raw_url: str) -> str:
    u = safe_str(raw_url)
    if not u:
        return ""
    try:
        parts = urlsplit(u if "://" in u else "https://" + u)
        host = (parts.netloc or "").lower()
        if host.startswith("www."):
            host = host[4:]
        if not host.endswith("bing.com") or not parts.path.startswith("/ck/a"):
            return u

        qs = dict(parse_qsl(parts.query, keep_blank_values=True))
        enc = safe_str(qs.get("u"))
        if not enc.startswith("a1"):
            return ""
        b64 = enc[2:]
        pad = "=" * ((4 - (len(b64) % 4)) % 4)
        try:
            decoded = base64.urlsafe_b64decode((b64 + pad).encode("utf-8")).decode("utf-8", errors="replace")
        except Exception:
            return ""
        return decoded if decoded.startswith("http") else ""
    except Exception:
        return ""


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Path to geo_updated.xlsx")
    ap.add_argument("--overwrite", action="store_true", help="Overwrite existing non-empty cells when merging duplicates")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx)
    ws_urls = wb["urls"]
    ws_bing = wb["bing_results"]
    h_urls = sheet_header_map(ws_urls)
    h_bing = sheet_header_map(ws_bing)

    if "url" not in h_urls:
        raise SystemExit("urls sheet missing url column")
    if "url" not in h_bing:
        raise SystemExit("bing_results sheet missing url column")

    # ---- decode ck/a in bing_results.url (in place); if undecodable, mark as bing_internal and blank url ----
    bing_url_col = h_bing["url"]
    page_type_col = h_bing.get("page_type")
    bing_ck = 0
    bing_decoded = 0
    bing_blank = 0
    for r in range(2, ws_bing.max_row + 1):
        v = ws_bing.cell(r, bing_url_col).value
        if not v:
            continue
        s = str(v).strip()
        if "bing.com/ck/a" in s.lower():
            bing_ck += 1
            dec = decode_bing_ck_url(s)
            if dec:
                ws_bing.cell(r, bing_url_col).value = dec
                bing_decoded += 1
            else:
                # This is usually a Bing-internal redirect (e.g., videos module).
                # Better to blank than to treat as a real page URL in analysis.
                ws_bing.cell(r, bing_url_col).value = None
                if page_type_col:
                    cur = ws_bing.cell(r, page_type_col).value
                    if cur is None or str(cur).strip() == "" or str(cur).strip().lower() == "nan":
                        ws_bing.cell(r, page_type_col).value = "bing_internal"
                bing_blank += 1

    # ---- decode ck/a in urls.url; if undecodable, clear the entire row ----
    url_col = h_urls["url"]
    url_rows: List[int] = []
    urls_ck = 0
    urls_decoded = 0
    urls_cleared = 0
    for r in range(2, ws_urls.max_row + 1):
        v = ws_urls.cell(r, url_col).value
        if not v:
            continue
        s = str(v).strip()
        if not s:
            continue
        if "bing.com/ck/a" in s.lower():
            urls_ck += 1
            dec = decode_bing_ck_url(s)
            if dec:
                ws_urls.cell(r, url_col).value = dec
                s = dec
                urls_decoded += 1
            else:
                # Not a real destination URL; clear row to prevent pollution/duplicates.
                for col in h_urls.values():
                    ws_urls.cell(r, col).value = None
                urls_cleared += 1
                continue
        url_rows.append(r)

    groups: Dict[str, List[int]] = {}
    for r in url_rows:
        s = str(ws_urls.cell(r, url_col).value).strip()
        k = normalize_url_key(s)
        if not k:
            continue
        groups.setdefault(k, []).append(r)

    dup_groups = {k: rows for k, rows in groups.items() if len(rows) > 1}

    # ---- merge duplicates: keep first row, fill from others, clear others ----
    merged = 0
    cleared = 0
    cols = list(h_urls.values())
    for _, rows in dup_groups.items():
        keep = rows[0]
        for r in rows[1:]:
            # merge cells
            for col in cols:
                if col == url_col:
                    continue
                src = ws_urls.cell(r, col).value
                if src is None or str(src).strip() == "":
                    continue
                dst = ws_urls.cell(keep, col).value
                if args.overwrite or dst is None or str(dst).strip() == "":
                    ws_urls.cell(keep, col).value = src
            # clear duplicate row
            for col in cols:
                ws_urls.cell(r, col).value = None
            cleared += 1
        merged += 1

    wb.save(args.xlsx)

    print("Cleanup done.")
    print(f"bing_results ck/a found: {bing_ck}, decoded: {bing_decoded}, blanked: {bing_blank}")
    print(f"urls ck/a found: {urls_ck}, decoded: {urls_decoded}, cleared: {urls_cleared}")
    print(f"urls duplicate groups merged: {merged}, rows cleared: {cleared}")


if __name__ == "__main__":
    main()

