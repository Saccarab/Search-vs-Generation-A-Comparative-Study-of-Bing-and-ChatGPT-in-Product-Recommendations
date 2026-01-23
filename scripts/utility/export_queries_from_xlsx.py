import openpyxl
import pandas as pd
from pathlib import Path

def export_queries_from_xlsx(xlsx_path, output_csv):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["runs"]
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    
    rid_idx = header.index("run_id") + 1
    query_idx = header.index("rewritten_query") + 1
    
    rows = []
    for r in range(2, ws.max_row + 1):
        rid = str(ws.cell(r, rid_idx).value or "").strip()
        query = str(ws.cell(r, query_idx).value or "").strip()
        
        if query and query.lower() != 'nan':
            rows.append({"run_id": rid, "query": query})
            
    df = pd.DataFrame(rows)
    df.to_csv(output_csv, index=False)
    print(f"Successfully exported {len(df)} queries to {output_csv}")

if __name__ == "__main__":
    export_queries_from_xlsx("geo-fresh.xlsx", "data/ingest/rewritten_queries.csv")
