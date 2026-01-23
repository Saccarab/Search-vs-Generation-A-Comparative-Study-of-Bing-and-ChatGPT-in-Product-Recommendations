import openpyxl
import pandas as pd
from pathlib import Path

def cleanup_orphan_urls(xlsx_path):
    if not Path(xlsx_path).exists():
        print(f"File not found: {xlsx_path}")
        return

    print(f"Analyzing {xlsx_path} for orphan URLs...")
    
    # Load all sheets to find used URLs
    wb = openpyxl.load_workbook(xlsx_path)
    
    used_urls = set()
    
    # 1. Collect URLs from citations
    if "citations" in wb.sheetnames:
        ws_cit = wb["citations"]
        h_cit = {str(ws_cit.cell(1, c).value).strip(): c for c in range(1, ws_cit.max_column + 1)}
        if "url" in h_cit:
            u_idx = h_cit["url"]
            for r in range(2, ws_cit.max_row + 1):
                val = ws_cit.cell(r, u_idx).value
                if val: used_urls.add(str(val).strip())
        print(f"  - Collected {len(used_urls)} unique URLs from 'citations'.")

    # 2. Collect URLs from bing_results
    if "bing_results" in wb.sheetnames:
        ws_bing = wb["bing_results"]
        h_bing = {str(ws_bing.cell(1, c).value).strip(): c for c in range(1, ws_bing.max_column + 1)}
        if "url" in h_bing:
            u_idx = h_bing["url"]
            prev_count = len(used_urls)
            for r in range(2, ws_bing.max_row + 1):
                val = ws_bing.cell(r, u_idx).value
                if val: used_urls.add(str(val).strip())
            print(f"  - Collected {len(used_urls) - prev_count} additional unique URLs from 'bing_results'.")

    # 3. Find and delete orphans from 'urls' sheet
    if "urls" in wb.sheetnames:
        ws_urls = wb["urls"]
        h_urls = {str(ws_urls.cell(1, c).value).strip(): c for c in range(1, ws_urls.max_column + 1)}
        if "url" not in h_urls:
            print("  - 'url' column not found in 'urls' sheet.")
            return
            
        u_idx = h_urls["url"]
        rows_to_delete = []
        orphan_samples = []
        
        for r in range(2, ws_urls.max_row + 1):
            url = str(ws_urls.cell(r, u_idx).value or "").strip()
            if not url: continue
            if url not in used_urls:
                rows_to_delete.append(r)
                if len(orphan_samples) < 5:
                    orphan_samples.append(url)
        
        if rows_to_delete:
            print(f"  - Found {len(rows_to_delete)} orphan URLs. Sample: {orphan_samples}")
            # Delete from bottom up to maintain indices
            for r in reversed(rows_to_delete):
                ws_urls.delete_rows(r)
            
            wb.save(xlsx_path)
            print(f"  - Successfully deleted {len(rows_to_delete)} orphans and saved {xlsx_path}.")
        else:
            print("  - No orphan URLs found.")

if __name__ == "__main__":
    cleanup_orphan_urls("geo-fresh.xlsx")
