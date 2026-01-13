"""
One-shot debug tool:
  - For a small list of URLs, ensure we have content (fetch if needed)
  - Then call Gemini with prompts/page_label_prompt_v1.txt and print what happens

This is meant for debugging failures from data/llm/page_labels_gemini.jsonl.

It does NOT send the workbook to Gemini. It only sends:
  URL + title/snippet (best-effort) + extracted text (truncated).

Usage (Windows):
  python -u scripts/llm/fetch_and_label_urls_debug.py ^
    --xlsx geo_updated.xlsx ^
    --content-root "C:\\Users\\User\\Documents\\thesis\\python_content\\debug_fetch_label" ^
    --urls "https://blog.openl.io/best-speech-translator/,https://voice-ping.com/blog/en-voice-translation-apps/"
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import os
import re
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Optional, Tuple
from urllib.parse import urlparse

import openpyxl
import requests


GEMINI_ENDPOINT_TMPL = "https://generativelanguage.googleapis.com/v1beta/{model_path}:generateContent?key={key}"

RE_SCRIPT_STYLE = re.compile(r"(?is)<(script|style|noscript)[^>]*>.*?</\\1>")
RE_TAGS = re.compile(r"<[^>]+>")
RE_SPACE = re.compile(r"\\s+")
WORD_RE = re.compile(r"[A-Za-z0-9]+(?:'[A-Za-z0-9]+)?")


def safe_str(v: object) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() == "nan" else s


def utc_now_iso() -> str:
    return dt.datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def short_hash(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()[:16]


def extract_domain(url: str) -> str:
    try:
        u = url.strip()
        if "://" not in u:
            u = "https://" + u
        host = (urlparse(u).netloc or "").lower()
        return host[4:] if host.startswith("www.") else host
    except Exception:
        return ""


def strip_html_to_text(html: str) -> str:
    if not html:
        return ""
    html2 = RE_SCRIPT_STYLE.sub(" ", html)
    txt = RE_TAGS.sub(" ", html2)
    return RE_SPACE.sub(" ", txt).strip()


def count_words(text: str) -> int:
    if not text:
        return 0
    return len(WORD_RE.findall(text))


def has_schema_markup_from_html(html: str) -> bool:
    h = (html or "").lower()
    if 'type="application/ld+json"' in h:
        return True
    if "itemscope" in h and "itemtype" in h:
        return True
    return False


def sheet_header_map(ws) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is None:
            continue
        k = str(v).strip()
        if k:
            m[k] = c
    return m


def find_row_by_key(ws, col_idx: int, key_val: str) -> Optional[int]:
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, col_idx).value
        if v is None:
            continue
        if str(v).strip() == key_val:
            return r
    return None


def append_row_first_truly_empty(ws, header: Dict[str, int], row: Dict[str, object]) -> int:
    for rr in range(2, ws.max_row + 1):
        has_any = False
        for col_idx in header.values():
            v = ws.cell(rr, col_idx).value
            if v is not None and str(v).strip() != "":
                has_any = True
                break
        if not has_any:
            for k, v in row.items():
                if k in header:
                    ws.cell(rr, header[k]).value = v
            return rr
    rr = ws.max_row + 1
    for k, v in row.items():
        if k in header:
            ws.cell(rr, header[k]).value = v
    return rr


def upsert_row(ws, header: Dict[str, int], key_col: str, key_val: str, row: Dict[str, object], overwrite: bool) -> int:
    key_idx = header[key_col]
    existing = find_row_by_key(ws, key_idx, key_val)
    if existing is None:
        row2 = dict(row)
        row2[key_col] = key_val
        return append_row_first_truly_empty(ws, header, row2)
    for k, v in row.items():
        if k not in header:
            continue
        if v is None:
            continue
        if isinstance(v, str) and v.strip() == "":
            continue
        if not overwrite:
            cur = ws.cell(existing, header[k]).value
            if cur is not None and str(cur).strip() != "":
                continue
        ws.cell(existing, header[k]).value = v
    return existing


def _gemini_model_path(model: str) -> str:
    m = (model or "").strip()
    if not m:
        return "models/gemini-flash-latest"
    return m if m.startswith("models/") else f"models/{m}"


def build_best_title_snippet(wb) -> Dict[str, Tuple[str, str]]:
    out: Dict[str, Tuple[str, str]] = {}
    if "bing_results" in wb.sheetnames:
        ws = wb["bing_results"]
        h = sheet_header_map(ws)
        if all(k in h for k in ("url", "result_title", "snippet")):
            for r in range(2, ws.max_row + 1):
                u = safe_str(ws.cell(r, h["url"]).value)
                if not u or u in out:
                    continue
                t = safe_str(ws.cell(r, h["result_title"]).value)
                s = safe_str(ws.cell(r, h["snippet"]).value)
                if t or s:
                    out[u] = (t, s)
    if "citations" in wb.sheetnames:
        ws = wb["citations"]
        h = sheet_header_map(ws)
        if all(k in h for k in ("url", "citation_title")):
            for r in range(2, ws.max_row + 1):
                u = safe_str(ws.cell(r, h["url"]).value)
                if not u or u in out:
                    continue
                t = safe_str(ws.cell(r, h["citation_title"]).value)
                if t:
                    out[u] = (t, "")
    return out


def fill_prompt(template: str, *, url: str, title: str, snippet: str, text: str) -> str:
    dom = extract_domain(url)
    return (
        template.replace("{URL}", url)
        .replace("{URL_DOMAIN}", dom)
        .replace("{TITLE}", title or "")
        .replace("{SNIPPET_OR_META_DESCRIPTION}", snippet or "")
        .replace("{EXTRACTED_TEXT}", text or "")
    )


@dataclass
class FetchResult:
    html: str
    final_url: str
    status: int
    error: str


def fetch_html(url: str, timeout: int) -> FetchResult:
    try:
        headers = {"User-Agent": "Mozilla/5.0 (compatible; GEOThesisBot/0.1)"}
        r = requests.get(url, headers=headers, timeout=timeout, allow_redirects=True)
        html = r.text if r.ok else ""
        return FetchResult(html=html, final_url=r.url, status=r.status_code, error="")
    except Exception as e:
        return FetchResult(html="", final_url=url, status=0, error=str(e))


def gemini_call_raw(*, api_key: str, model: str, prompt: str, timeout: int) -> Tuple[int, str]:
    payload = {"contents": [{"role": "user", "parts": [{"text": prompt}]}], "generationConfig": {"temperature": 0.0}}
    ep = GEMINI_ENDPOINT_TMPL.format(model_path=_gemini_model_path(model), key=api_key)
    r = requests.post(ep, json=payload, timeout=timeout)
    return r.status_code, (r.text or "")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="geo_updated.xlsx")
    ap.add_argument("--prompt-file", default="prompts/page_label_prompt_v1.txt")
    ap.add_argument("--content-root", required=True)
    ap.add_argument("--urls", required=True, help="Comma-separated URLs")
    ap.add_argument("--model", default="gemini-3-flash-preview")
    ap.add_argument("--fetch-timeout", type=int, default=25)
    ap.add_argument("--gemini-timeout", type=int, default=25)
    ap.add_argument("--text-max-chars", type=int, default=4000, help="Max extracted text chars to send to Gemini (default: 4000)")
    ap.add_argument("--overwrite-content", action="store_true")
    ap.add_argument("--overwrite-labels", action="store_true")
    ap.add_argument("--sleep", type=float, default=0.0)
    ap.add_argument("--log-jsonl", default="data/llm/debug_fetch_label.jsonl")
    args = ap.parse_args()

    api_key = (
        os.getenv("GEMINI_API_KEY", "").strip()
        or os.getenv("GOOGLE_API_KEY", "").strip()
        or os.getenv("GOOGLE_GENERATIVE_AI_API_KEY", "").strip()
    )
    if not api_key:
        raise SystemExit("Missing GEMINI_API_KEY in env.")

    wb = openpyxl.load_workbook(args.xlsx)
    ws_urls = wb["urls"]
    ws_listicles = wb["listicles"]
    ws_products = wb["listicle_products"]
    hu = sheet_header_map(ws_urls)
    hl = sheet_header_map(ws_listicles)
    hp = sheet_header_map(ws_products)

    required_urls = ["url", "domain", "content_path", "content_word_count", "has_schema_markup", "fetched_at", "type"]
    miss = [c for c in required_urls if c not in hu]
    if miss:
        raise SystemExit(f"urls sheet missing columns: {miss}")

    template = Path(args.prompt_file).read_text(encoding="utf-8")
    lookup = build_best_title_snippet(wb)
    out_dir = Path(args.content_root)
    out_dir.mkdir(parents=True, exist_ok=True)
    os.makedirs(Path(args.log_jsonl).parent, exist_ok=True)

    urls = [safe_str(u) for u in args.urls.split(",") if safe_str(u)]
    for url in urls:
        print("\n=== URL ===", flush=True)
        print(url, flush=True)

        # locate/create urls row
        r = find_row_by_key(ws_urls, hu["url"], url)
        if r is None:
            r = append_row_first_truly_empty(ws_urls, hu, {"url": url, "domain": extract_domain(url)})

        cp = safe_str(ws_urls.cell(r, hu["content_path"]).value)
        has_file = bool(cp and os.path.exists(cp))
        if (not has_file) or args.overwrite_content:
            print(f"Fetching content (timeout={args.fetch_timeout}s)...", flush=True)
            fr = fetch_html(url, timeout=args.fetch_timeout)
            if fr.status != 200 or not fr.html:
                print("Fetch failed:", fr.status, fr.error, flush=True)
                with open(args.log_jsonl, "a", encoding="utf-8") as lf:
                    lf.write(json.dumps({"url": url, "phase": "fetch", "ok": False, "status": fr.status, "error": fr.error}) + "\n")
                continue
            text = strip_html_to_text(fr.html)
            if len(text) < 200:
                print("Extracted text too short:", len(text), flush=True)
                with open(args.log_jsonl, "a", encoding="utf-8") as lf:
                    lf.write(json.dumps({"url": url, "phase": "fetch", "ok": False, "status": fr.status, "error": "text too short"}) + "\n")
                continue
            fname = f"{short_hash(url)}.txt"
            fpath = out_dir / fname
            fpath.write_text(text, encoding="utf-8", errors="ignore")
            ws_urls.cell(r, hu["content_path"]).value = str(fpath)
            ws_urls.cell(r, hu["content_word_count"]).value = int(count_words(text))
            ws_urls.cell(r, hu["has_schema_markup"]).value = 1 if has_schema_markup_from_html(fr.html) else 0
            ws_urls.cell(r, hu["domain"]).value = ws_urls.cell(r, hu["domain"]).value or extract_domain(url)
            ws_urls.cell(r, hu["fetched_at"]).value = utc_now_iso()
            wb.save(args.xlsx)
            print("Saved content to:", str(fpath), flush=True)
        else:
            print("Using existing content_path:", cp, flush=True)

        cp2 = safe_str(ws_urls.cell(r, hu["content_path"]).value)
        if not cp2 or not os.path.exists(cp2):
            print("No content file available, cannot label.", flush=True)
            continue
        text = Path(cp2).read_text(encoding="utf-8", errors="ignore")
        text_send = text[: max(0, int(args.text_max_chars or 0))] if args.text_max_chars else ""
        title, snippet = lookup.get(url, ("", ""))
        prompt = fill_prompt(template, url=url, title=title, snippet=snippet, text=text_send)
        print(f"Prompt chars: {len(prompt)} | text_sent_chars: {len(text_send)}", flush=True)

        print(f"Calling Gemini (timeout={args.gemini_timeout}s)...", flush=True)
        t0 = time.time()
        try:
            status, body = gemini_call_raw(api_key=api_key, model=args.model, prompt=prompt, timeout=args.gemini_timeout)
            dt_s = time.time() - t0
            print("Gemini HTTP:", status, f"({dt_s:.1f}s)", flush=True)
            print("Response snippet:", body[:800], flush=True)
            if status != 200:
                with open(args.log_jsonl, "a", encoding="utf-8") as lf:
                    lf.write(json.dumps({"url": url, "phase": "gemini", "ok": False, "http_status": status, "http_body_snippet": body[:800]}) + "\n")
                continue

            data = json.loads(body)
            out_text = data["candidates"][0]["content"]["parts"][0].get("text", "")
            out_text = re.sub(r"^```(?:json)?\\s*|\\s*```$", "", (out_text or "").strip(), flags=re.I | re.M)
            parsed = json.loads(out_text)
            print("Parsed top-level keys:", list(parsed.keys()) if isinstance(parsed, dict) else type(parsed), flush=True)

            # Write back minimally (urls + listicles + products), optional overwrite for labels
            uo = parsed.get("urls") if isinstance(parsed, dict) else None
            if isinstance(uo, dict):
                upsert_row(ws_urls, hu, "url", url, uo, overwrite=args.overwrite_labels)
            lo = parsed.get("listicles") if isinstance(parsed, dict) else None
            if isinstance(lo, dict):
                upsert_row(ws_listicles, hl, "listicle_url", url, lo, overwrite=args.overwrite_labels)
            po = parsed.get("listicle_products") if isinstance(parsed, dict) else None
            if isinstance(po, list):
                for item in po:
                    if not isinstance(item, dict):
                        continue
                    # Ensure parent key present
                    item["listicle_url"] = safe_str(item.get("listicle_url") or url)
                    append_row_first_truly_empty(ws_products, hp, item)

            wb.save(args.xlsx)
            with open(args.log_jsonl, "a", encoding="utf-8") as lf:
                lf.write(json.dumps({"url": url, "phase": "gemini", "ok": True, "response": parsed}) + "\n")
        except Exception as e:
            dt_s = time.time() - t0
            print("Gemini error:", type(e).__name__, str(e)[:600], f"({dt_s:.1f}s)", flush=True)
            with open(args.log_jsonl, "a", encoding="utf-8") as lf:
                lf.write(json.dumps({"url": url, "phase": "gemini", "ok": False, "error": {"type": type(e).__name__, "message": str(e)[:1200]}}) + "\n")

        if args.sleep:
            time.sleep(args.sleep)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

