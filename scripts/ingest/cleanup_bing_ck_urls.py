"""
Replace Bing click-tracking URLs (bing.com/ck/a?...) stored in geo_updated.xlsx with their decoded destinations.

Targets:
  - urls.url (+ update urls.domain)
  - bing_results.url (+ update bing_results.result_domain)

If a ck/a URL cannot be decoded to an http(s) URL, we blank it (so it can be reviewed/removed).
"""

from __future__ import annotations

import argparse
import base64
from pathlib import Path
from urllib.parse import parse_qsl, urlsplit

import openpyxl


def extract_domain(raw_url: str) -> str:
    if not raw_url or not isinstance(raw_url, str):
        return ""
    url = raw_url.strip()
    if not url:
        return ""
    if "://" not in url:
        url = "https://" + url
    parts = urlsplit(url)
    host = (parts.netloc or "").lower()
    if host.startswith("www."):
        host = host[4:]
    return host


def decode_bing_ck_url(raw_url: str) -> str:
    url = (raw_url or "").strip()
    if not url:
        return ""
    parts = urlsplit(url if "://" in url else "https://" + url)
    host = (parts.netloc or "").lower()
    if host.startswith("www."):
        host = host[4:]
    if not host.endswith("bing.com") or not parts.path.startswith("/ck/a"):
        return url
    qs = dict(parse_qsl(parts.query, keep_blank_values=True))
    u = (qs.get("u") or "").strip()
    if not u.startswith("a1"):
        return ""
    b64 = u[2:]
    pad = "=" * ((4 - (len(b64) % 4)) % 4)
    try:
        decoded = base64.urlsafe_b64decode((b64 + pad).encode("utf-8")).decode("utf-8", errors="replace")
    except Exception:
        return ""
    return decoded if decoded.startswith("http") else ""


def header_map(ws):
    return {str(ws.cell(1, c).value).strip(): c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Path to geo_updated.xlsx")
    args = ap.parse_args()

    xlsx = Path(args.xlsx)
    wb = openpyxl.load_workbook(xlsx)

    changed = 0
    blanked = 0

    # urls sheet
    if "urls" in wb.sheetnames:
        ws = wb["urls"]
        h = header_map(ws)
        url_c = h.get("url")
        dom_c = h.get("domain")
        if url_c:
            for r in range(2, ws.max_row + 1):
                v = ws.cell(r, url_c).value
                if not (isinstance(v, str) and "bing.com/ck/a" in v):
                    continue
                decoded = decode_bing_ck_url(v)
                if decoded:
                    ws.cell(r, url_c).value = decoded
                    if dom_c:
                        ws.cell(r, dom_c).value = extract_domain(decoded)
                    changed += 1
                else:
                    ws.cell(r, url_c).value = None
                    if dom_c:
                        ws.cell(r, dom_c).value = None
                    blanked += 1

    # bing_results sheet
    if "bing_results" in wb.sheetnames:
        ws = wb["bing_results"]
        h = header_map(ws)
        url_c = h.get("url")
        dom_c = h.get("result_domain")
        if url_c:
            for r in range(2, ws.max_row + 1):
                v = ws.cell(r, url_c).value
                if not (isinstance(v, str) and "bing.com/ck/a" in v):
                    continue
                decoded = decode_bing_ck_url(v)
                if decoded:
                    ws.cell(r, url_c).value = decoded
                    if dom_c:
                        ws.cell(r, dom_c).value = extract_domain(decoded)
                    changed += 1
                else:
                    ws.cell(r, url_c).value = None
                    if dom_c:
                        ws.cell(r, dom_c).value = None
                    blanked += 1

    wb.save(xlsx)
    print(f"Updated {xlsx}")
    print(f"Decoded+replaced: {changed}")
    print(f"Blanked (couldn't decode): {blanked}")


if __name__ == "__main__":
    main()


