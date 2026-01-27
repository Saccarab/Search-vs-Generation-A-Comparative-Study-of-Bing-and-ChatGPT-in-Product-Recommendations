import openpyxl
import pandas as pd
from pathlib import Path

def find_missing_rewritten_queries(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["runs"]
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    
    try:
        rid_idx = header.index("run_id") + 1
        query_idx = header.index("rewritten_query") + 1
        cit_idx = header.index("citation_count") + 1
    except ValueError as e:
        print(f"Error: Missing column in header: {e}")
        return

    missing = []
    for r in range(2, ws.max_row + 1):
        rid = str(ws.cell(r, rid_idx).value or "").strip()
        query = str(ws.cell(r, query_idx).value or "").strip()
        cit_count = ws.cell(r, cit_idx).value
        
        # citation_count can be 0 or None if not ingested correctly, 
        # but the user specifically asked for those where it exists but query is missing.
        try:
            cit_count = int(cit_count) if cit_count is not None else 0
        except:
            cit_count = 0
            
        # Check if query is missing (empty or literal 'nan')
        if not query or query.lower() == 'nan':
            missing.append({
                "run_id": rid,
                "citation_count": cit_count
            })
            
    return missing

if __name__ == "__main__":
    xlsx = "geo-fresh.xlsx"
    missing = find_missing_rewritten_queries(xlsx)
    
    if missing:
        # Filter for those with citations (likely search triggered but not captured)
        likely_failures = [m for m in missing if m['citation_count'] > 0]
        
        print(f"Found {len(missing)} runs with missing rewritten_query.")
        print(f"Of those, {len(likely_failures)} have citation_count > 0 (High Priority):")
        for m in likely_failures:
            print(f"  - {m['run_id']} (Citations: {m['citation_count']})")
            
        print("\nRuns with 0 citations and no query (Likely no search triggered):")
        no_search = [m for m in missing if m['citation_count'] == 0]
        for m in no_search[:10]: # Show first 10
             print(f"  - {m['run_id']}")
        if len(no_search) > 10:
            print(f"  ... and {len(no_search) - 10} more.")
    else:
        print("No missing rewritten_query found.")
