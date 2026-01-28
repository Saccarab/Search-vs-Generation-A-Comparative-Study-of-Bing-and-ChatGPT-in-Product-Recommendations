import argparse
import datetime as dt
import json
import base64
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from urllib.parse import urlsplit
import openpyxl
import pandas as pd

# Reuse core functions from the original ingest script
def utc_now_iso() -> str:
    return dt.datetime.utcnow().replace(microsecond=0).isoformat() + "Z"

def safe_str(v: object) -> str:
    if v is None: return ""
    try:
        if pd.isna(v): return ""
    except: pass
    s = str(v).strip()
    return "" if s.lower() == "nan" else s

def extract_domain(raw_url: str) -> str:
    if not raw_url: return ""
    url = str(raw_url).strip()
    if "://" not in url: url = "https://" + url
    try:
        host = (urlsplit(url).netloc or "").lower()
        if host.startswith("www."): host = host[4:]
        return host
    except: return ""

def parse_sources_json(cell: object) -> List[Dict[str, str]]:
    if cell is None: return []
    s = str(cell).strip()
    if not s or s == "[]": return []
    try:
        arr = json.loads(s)
    except: return []
    out = []
    if isinstance(arr, list):
        for it in arr:
            if isinstance(it, dict) and isinstance(it.get("url"), str):
                url = it["url"].strip()
                out.append({
                    "url": url,
                    "title": str(it.get("title") or "").strip(),
                    "description": str(it.get("description") or "").strip(),
                    "domain": str(it.get("domain") or extract_domain(url)).strip(),
                })
            elif isinstance(it, str):
                url = it.strip()
                out.append({"url": url, "title": "", "description": "", "domain": extract_domain(url)})
    return out

def parse_items_json(cell: object) -> List[dict]:
    if cell is None: return []
    s = str(cell).strip()
    if not s or s == "[]": return []
    try:
        arr = json.loads(s)
        return arr if isinstance(arr, list) else []
    except: return []

def sheet_header_map(ws) -> Dict[str, int]:
    return {str(ws.cell(row=1, column=c).value).strip(): c for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}

def append_row(ws, header: Dict[str, int], row: Dict[str, object]) -> int:
    r = None
    for rr in range(2, ws.max_row + 2):
        has_any = False
        for col_idx in header.values():
            v = ws.cell(row=rr, column=col_idx).value
            if v is not None and str(v).strip() != "":
                has_any = True
                break
        if not has_any:
            r = rr
            break
    if r is None: r = ws.max_row + 1
    for k, v in row.items():
        if k in header: ws.cell(row=r, column=header[k]).value = v
    return r

def split_entity_names(item_name: str) -> List[str]:
    s = safe_str(item_name)
    if not s: return []
    for sep in [" & ", " and ", " / ", " | ", " + "]:
        if sep in s:
            parts = [p.strip() for p in s.split(sep) if p.strip()]
            return parts if len(parts) >= 2 else [s]
    if ", " in s:
        parts = [p.strip() for p in s.split(",") if p.strip()]
        return parts if len(parts) >= 2 else [s]
    return [s]

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--chatgpt", required=True)
    ap.add_argument("--user-location", default="US")
    ap.add_argument("--vpn-status", default="on")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx)
    ws_runs = wb["runs"]
    ws_cit = wb["citations"]
    ws_urls = wb["urls"]

    h_runs = sheet_header_map(ws_runs)
    h_cit = sheet_header_map(ws_cit)
    h_urls = sheet_header_map(ws_urls)

    # 1. Map current max run numbers per prompt
    rid_col = h_runs["run_id"]
    max_runs = {}
    for r in range(2, ws_runs.max_row + 1):
        rid = str(ws_runs.cell(r, rid_col).value or "").strip()
        if not rid or "_" not in rid: continue
        parts = rid.split("_r")
        if len(parts) == 2:
            pid, num = parts[0], int(parts[1])
            max_runs[pid] = max(max_runs.get(pid, 0), num)

    print(f"Current max runs detected: {max_runs}")

    # 2. Process patch CSV
    chat = pd.read_csv(args.chatgpt, engine="python")
    
    # Store existing URLs for dedupe in 'urls' sheet
    existing_urls = set()
    if "url" in h_urls:
        for r in range(2, ws_urls.max_row + 1):
            v = ws_urls.cell(r, h_urls["url"]).value
            if v: existing_urls.add(str(v).strip())

    for _, r in chat.iterrows():
        prompt_id = str(r["prompt_id"])
        # Increment run number
        next_num = max_runs.get(prompt_id, 0) + 1
        max_runs[prompt_id] = next_num
        run_id = f"{prompt_id}_r{next_num}"
        
        print(f"Ingesting {run_id}...")

        # --- runs ---
        rewritten = safe_str(r.get("generated_search_query", ""))
        cited = parse_sources_json(r.get("sources_cited_json"))
        citation_count = len({c["url"] for c in cited if c.get("url")})

        append_row(ws_runs, h_runs, {
            "run_id": run_id,
            "prompt_id": prompt_id,
            "run_ts_utc": utc_now_iso(),
            "citation_count": citation_count,
            "rewritten_query": rewritten,
            "user_location": args.user_location,
            "vpn_status": args.vpn_status,
        })

        # --- citations ---
        url_meta = {}
        sources_panel_pos = {}
        cited_arr = parse_sources_json(r.get("sources_cited_json"))
        additional_arr = parse_sources_json(r.get("sources_additional_json"))
        
        for i, it in enumerate(cited_arr, start=1):
            u = it.get("url", "").strip()
            if u:
                url_meta[u] = it
                if u not in sources_panel_pos: sources_panel_pos[u] = i
                append_row(ws_cit, h_cit, {
                    "run_id": run_id,
                    "cite_position": i,
                    "citation_title": it.get("title", ""),
                    "url": u,
                    "citation_domain": it.get("domain", "") or extract_domain(u),
                    "citation_type": "cited",
                    "capture_method": "sources_panel",
                    "notes": it.get("description", ""),
                    "citation_scope": "sources_panel",
                })

        for i, it in enumerate(additional_arr, start=1):
            u = it.get("url", "").strip()
            if u:
                url_meta[u] = it
                if u not in sources_panel_pos: sources_panel_pos[u] = i
                append_row(ws_cit, h_cit, {
                    "run_id": run_id,
                    "cite_position": i,
                    "citation_title": it.get("title", ""),
                    "url": u,
                    "citation_domain": it.get("domain", "") or extract_domain(u),
                    "citation_type": "additional",
                    "capture_method": "sources_panel",
                    "notes": it.get("description", ""),
                    "citation_scope": "sources_panel",
                })

        items = parse_items_json(r.get("items_json"))
        for item in items:
            item_name = safe_str(item.get("item_name", ""))
            item_text = safe_str(item.get("item_text", ""))
            item_pos = item.get("item_position")
            chip_groups = item.get("chip_groups", []) or []
            for g_idx, group in enumerate(chip_groups, start=1):
                links = group.get("links", []) if isinstance(group, dict) else []
                group_size = len(links)
                entity_names = split_entity_names(item_name)
                entity_name = entity_names[g_idx-1] if entity_names and 1 <= g_idx <= len(entity_names) else ""
                
                for j, link in enumerate(links, start=1):
                    url = safe_str(link)
                    if not url: continue
                    meta = url_meta.get(url, {})
                    append_row(ws_cit, h_cit, {
                        "run_id": run_id,
                        "cite_position": sources_panel_pos.get(url),
                        "citation_title": safe_str(meta.get("title")),
                        "url": url,
                        "citation_domain": safe_str(meta.get("domain")) or extract_domain(url),
                        "citation_type": "inline",
                        "capture_method": "inline_chips",
                        "citation_scope": "inline",
                        "item_position": item_pos,
                        "item_name": item_name,
                        "item_text": item_text,
                        "item_entity_name": entity_name,
                        "citation_group_index": g_idx,
                        "citation_group_size": group_size,
                        "citation_in_group_rank": j,
                    })

        # --- urls sheet updates ---
        all_urls_in_run = set(url_meta.keys())
        for u in all_urls_in_run:
            if u and u not in existing_urls:
                existing_urls.add(u)
                append_row(ws_urls, h_urls, {
                    "url": u,
                    "domain": extract_domain(u),
                })

    wb.save(args.xlsx)
    print(f"Successfully ingested patches into {args.xlsx}")

if __name__ == "__main__":
    main()
