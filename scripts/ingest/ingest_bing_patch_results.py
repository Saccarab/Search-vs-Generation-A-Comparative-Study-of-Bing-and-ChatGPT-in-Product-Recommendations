import argparse
import pandas as pd
import openpyxl
from pathlib import Path
from typing import Dict, Set, List
import base64
from urllib.parse import urlsplit, parse_qsl, urlunsplit, urlencode

def safe_str(v: object) -> str:
    if v is None: return ""
    try:
        if pd.isna(v): return ""
    except: pass
    s = str(v).strip()
    return "" if s.lower() == "nan" else s

def normalize_url_key(raw_url: str) -> str:
    u = safe_str(raw_url)
    if not u: return ""
    if "://" not in u: u = "https://" + u
    try:
        p = urlsplit(u)
        host = (p.netloc or "").lower()
        if host.startswith("www."): host = host[4:]
        path = p.path or "/"
        if len(path) > 1 and path.endswith("/"): path = path[:-1]
        # Drop tracking params
        kept = []
        for k, v in parse_qsl(p.query, keep_blank_values=True):
            lk = k.lower()
            if lk.startswith("utm_") or lk in {"gclid", "fbclid", "msclkid", "yclid", "ref"}:
                continue
            kept.append((k, v))
        q = urlencode(kept, doseq=True)
        return urlunsplit(("", host, path, q, "")).lstrip("/")
    except: return u

def decode_bing_ck_url(raw_url: str) -> str:
    u = safe_str(raw_url)
    if not u or "bing.com/ck/a" not in u.lower(): return u
    try:
        parts = urlsplit(u if "://" in u else "https://" + u)
        qs = dict(parse_qsl(parts.query, keep_blank_values=True))
        enc = safe_str(qs.get("u"))
        if not enc.startswith("a1"): return ""
        b64 = enc[2:]
        pad = "=" * ((4 - (len(b64) % 4)) % 4)
        decoded = base64.urlsafe_b64decode((b64 + pad).encode("utf-8")).decode("utf-8", errors="replace")
        return decoded if decoded.startswith("http") else ""
    except: return ""

def sheet_header_map(ws) -> Dict[str, int]:
    return {str(ws.cell(row=1, column=c).value).strip(): c for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}

def ingest_bing_patch(xlsx_path, bing_csv_path):
    print(f"Loading {bing_csv_path}...")
    df = pd.read_csv(bing_csv_path, engine="python")
    
    # Identify run_ids being updated
    patch_run_ids = set(df['run_id'].unique())
    print(f"Updating data for run_ids: {patch_run_ids}")

    wb = openpyxl.load_workbook(xlsx_path)
    
    # 1. Clean up old data for these run_ids from both sheets
    for sn in ["bing_results", "bing_results_raw"]:
        if sn not in wb.sheetnames: continue
        ws = wb[sn]
        h = sheet_header_map(ws)
        if "run_id" not in h: continue
        rid_col = h["run_id"]
        
        rows_to_delete = []
        for r in range(2, ws.max_row + 1):
            rid = str(ws.cell(r, rid_col).value or "").strip()
            if rid in patch_run_ids:
                rows_to_delete.append(r)
        
        for r in reversed(rows_to_delete):
            ws.delete_rows(r)
        print(f"  - {sn}: Deleted {len(rows_to_delete)} old rows.")

    ws_raw = wb["bing_results_raw"]
    ws_main = wb["bing_results"]
    h_raw = sheet_header_map(ws_raw)
    h_main = sheet_header_map(ws_main)
    
    # 2. Ingest into bing_results_raw (exactly as is, with URL decoding)
    now = pd.Timestamp.now().isoformat()
    for _, r in df.iterrows():
        url_raw = safe_str(r.get("url"))
        url = decode_bing_ck_url(url_raw)
        if not url: url = url_raw # fallback

        row_data = {
            "run_id": safe_str(r.get("run_id")),
            "bing_query": safe_str(r.get("query")),
            "result_rank": safe_str(r.get("position")),
            "result_title": safe_str(r.get("title")),
            "url": url,
            "raw_url": url_raw,
            "result_domain": safe_str(r.get("domain")),
            "snippet": safe_str(r.get("snippet")),
            "page_num": safe_str(r.get("page_num")),
            "captured_at": now
        }
        
        new_row = ws_raw.max_row + 1
        for k, v in row_data.items():
            if k in h_raw:
                ws_raw.cell(new_row, h_raw[k]).value = v

    # 3. Ingest into bing_results (deduplicated per run_id, capped at 30)
    # Process the dataframe to deduplicate URLs per run_id
    processed_rows = []
    for rid in patch_run_ids:
        run_df = df[df['run_id'] == rid].copy()
        run_df['decoded_url'] = run_df['url'].apply(decode_bing_ck_url)
        # Fix empty decoded urls
        run_df.loc[run_df['decoded_url'] == "", 'decoded_url'] = run_df['url']
        
        # Deduplicate by URL (keep first/lowest rank)
        run_df = run_df.drop_duplicates(subset=['decoded_url']).head(30)
        
        for _, r in run_df.iterrows():
            processed_rows.append({
                "run_id": rid,
                "bing_query": safe_str(r.get("query")),
                "result_rank": safe_str(r.get("position")),
                "result_title": safe_str(r.get("title")),
                "url": r['decoded_url'],
                "result_domain": safe_str(r.get("domain")),
                "snippet": safe_str(r.get("snippet")),
                "page_num": safe_str(r.get("page_num")),
                "captured_at": now
            })

    for row in processed_rows:
        new_row = ws_main.max_row + 1
        for k, v in row.items():
            if k in h_main:
                ws_main.cell(new_row, h_main[k]).value = v

    # 4. Update URLs sheet with any new URLs
    ws_urls = wb["urls"]
    h_urls = sheet_header_map(ws_urls)
    existing_urls = set()
    for r in range(2, ws_urls.max_row + 1):
        v = ws_urls.cell(r, h_urls["url"]).value
        if v: existing_urls.add(str(v).strip())
    
    new_urls_count = 0
    for row in processed_rows:
        u = row["url"]
        if u and u not in existing_urls:
            existing_urls.add(u)
            new_row = ws_urls.max_row + 1
            ws_urls.cell(new_row, h_urls["url"]).value = u
            if "domain" in h_urls:
                ws_urls.cell(new_row, h_urls["domain"]).value = row["result_domain"]
            new_urls_count += 1

    wb.save(xlsx_path)
    print(f"Successfully ingested {len(df)} raw rows and {len(processed_rows)} deduped rows into {xlsx_path}.")
    print(f"Added {new_urls_count} new unique URLs to the 'urls' sheet.")

if __name__ == "__main__":
    ingest_bing_patch("geo-fresh.xlsx", "c:/Users/User/Downloads/bing_results_2026-01-18T01-44-15.csv")
